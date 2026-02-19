from django.shortcuts import render, redirect, get_object_or_404
from django.contrib.auth.decorators import login_required
from collections import defaultdict
from django.contrib.auth import login, logout, authenticate, update_session_auth_hash
from django.contrib.auth.forms import UserCreationForm, PasswordChangeForm
from django.views.decorators.csrf import ensure_csrf_cookie
from django.contrib import messages
from django.core.paginator import Paginator
from django.db.models import Count, Avg, Q, Max
from django.http import JsonResponse, HttpResponse
from django.utils import timezone
from django.contrib.auth.forms import UserCreationForm
from django import forms
from datetime import timedelta
import csv
import re
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from datetime import datetime
from .models import Complaint, Store, User, Department, Escala, IndicadorDesempenho, ObservacaoDesempenho, Lista, Activity, AuditLog, StoreAudit, StoreAuditItem, StoreAuditIssue, MetaMensalGlobal, SystemNotification
from .forms import ComplaintForm, StoreForm


def login_view_custom(request):
    """View customizada de login que verifica se o usuário está ativo e aceita e-mail"""
    if request.user.is_authenticated:
        return redirect('dashboard')
    
    if request.method == 'POST':
        email = request.POST.get('email')
        password = request.POST.get('password')
        
        if email and password:
            # Buscar usuário por e-mail
            try:
                user = User.objects.get(email=email)
                # Autenticar usando o username do usuário encontrado
                user = authenticate(request, username=user.username, password=password)
                
                if user is not None:
                    if user.ativo:
                        login(request, user)
                        # Log acess history
                        try:
                            AuditLog.objects.create(
                                usuario=user,
                                action='login',
                                target_type='User',
                                target_id=user.id,
                                detalhes_json={'ip': request.META.get('REMOTE_ADDR'), 'user_agent': request.META.get('HTTP_USER_AGENT')}
                            )
                        except Exception as e:
                            print(f"Error logging login: {e}")
                            
                        messages.success(request, f'Bem-vindo, {user.get_full_name() or user.username}!')
                        next_url = request.GET.get('next', '/')
                        return redirect(next_url)
                    else:
                        messages.error(request, 'Sua conta está inativa. Entre em contato com o administrador.')
                else:
                    messages.error(request, 'E-mail ou senha incorretos. Tente novamente.')
            except User.DoesNotExist:
                messages.error(request, 'E-mail ou senha incorretos. Tente novamente.')
            except User.MultipleObjectsReturned:
                # Se houver múltiplos usuários com o mesmo e-mail, tentar autenticar com o primeiro
                user = User.objects.filter(email=email).first()
                user = authenticate(request, username=user.username, password=password)
                if user is not None and user.ativo:
                    login(request, user)
                    # Log acess history
                    try:
                        AuditLog.objects.create(
                            usuario=user,
                            action='login',
                            target_type='User',
                            target_id=user.id,
                            detalhes_json={'ip': request.META.get('REMOTE_ADDR'), 'user_agent': request.META.get('HTTP_USER_AGENT')}
                        )
                    except Exception as e:
                        print(f"Error logging login: {e}")

                    messages.success(request, f'Bem-vindo, {user.get_full_name() or user.username}!')
                    next_url = request.GET.get('next', '/')
                    return redirect(next_url)
                else:
                    messages.error(request, 'E-mail ou senha incorretos. Tente novamente.')
        else:
            messages.error(request, 'Por favor, preencha todos os campos.')
    
    return render(request, 'core/login.html')


@login_required
def change_department(request, dept_id):
    if not request.user.is_administrador():
        return redirect('dashboard')
    
    # dept_id == 0 (Global) removido conforme solicitação
    if dept_id == 0:
        return redirect('dashboard')

    request.session['selected_department_id'] = dept_id
    from .models import Department
    dept = get_object_or_404(Department, id=dept_id)
    messages.success(request, f"Departamento alterado para: {dept.name}")
    
    # Redirecionar para a página principal de cada departamento
    if dept.name == 'NRS Suporte':
        return redirect('escala')
    elif dept.name == 'CS Clientes':
        return redirect('dashboard')
    elif dept.name == 'Onboarding':
        return redirect('onboarding_dev_1')
    
    return redirect('dashboard')

@login_required
def dashboard(request):
    # Filtro por departamento
    selected_dept_id = request.session.get('selected_department_id')
    
    if not request.user.is_administrador():
        # Redirecionar usuários do NRS Suporte para a escala (página operacional principal deles)
        if request.user.department and request.user.department.name == 'NRS Suporte':
            return redirect('escala')
        queryset = Complaint.objects.filter(department=request.user.department)
    else:
        # Se houver depto na sessão, verificar se é o NRS Suporte para redirecionar
        if selected_dept_id:
            from .models import Department
            current_dept = Department.objects.filter(id=selected_dept_id).first()
            if current_dept and current_dept.name == 'NRS Suporte':
                return redirect('escala')
            queryset = Complaint.objects.filter(department_id=selected_dept_id)
        else:
            # Se é admin mas não tem depto na sessão (primeiro acesso), 
            # tenta buscar o NRS Suporte e redirecionar
            from .models import Department
            nrs_dept = Department.objects.filter(name='NRS Suporte').first()
            if nrs_dept:
                request.session['selected_department_id'] = nrs_dept.id
                return redirect('escala')
            queryset = Complaint.objects.all()
        
    # OTIMIZAÇÃO: Usar aggregate para buscar contadores em uma única query
    stats = queryset.aggregate(
        total=Count('id'),
        pending=Count('id', filter=Q(status='pendente')),
        em_replica=Count('id', filter=Q(status='em_replica')),
        resolved=Count('id', filter=Q(status='resolvido')),
        awaiting=Count('id', filter=Q(status='aguardando_avaliacao')),
        em_andamento=Count('id', filter=Q(status='em_andamento'))
    )

    total_complaints = stats['total']
    pending = stats['pending']
    em_replica = stats['em_replica']
    resolved = stats['resolved']
    awaiting = stats['awaiting']
    em_andamento = stats['em_andamento']
    
    # Ranking de lojas com mais reclamações (top 5)
    top_stores_ranking = queryset.values('loja_cod').annotate(
        count=Count('id')
    ).order_by('-count')[:5]
    
    # Gráficos
    # Gráficos - Otimizado
    days = 30
    date_threshold = timezone.now().date() - timedelta(days=days)
    
    # Query única para agrupar por data
    daily_counts = queryset.filter(
        created_at__date__gte=date_threshold
    ).values('created_at__date').annotate(count=Count('id'))
    
    # Transformar em dicionário para lookup rápido
    counts_map = {item['created_at__date']: item['count'] for item in daily_counts}
    
    complaints_by_period = []
    for i in range(days):
        date = timezone.now().date() - timedelta(days=days-1-i) # Order from oldest to newest usually? Current code was 29-i (descending?)
        # Current code: for i in range(30): date = ... - (29-i). This goes from Oldest to Newest.
        # i=0: -29 days. i=29: -0 days.
        count = counts_map.get(date, 0)
        complaints_by_period.append({'day': date.isoformat(), 'count': count})
    
    top_stores = queryset.values('loja_cod').annotate(
        count=Count('id')
    ).order_by('-count')[:10]
    
    satisfaction_by_store = queryset.filter(
        nota_satisfacao__isnull=False
    ).values('loja_cod').annotate(
        avg=Avg('nota_satisfacao')
    )
    
    complaints_by_status = queryset.values('status').annotate(
        count=Count('id')
    )
    
    recent_complaints = queryset.select_related('analista').order_by('-created_at')[:10]
    
    # Estatísticas adicionais
    avg_satisfaction = queryset.filter(nota_satisfacao__isnull=False).aggregate(avg=Avg('nota_satisfacao'))['avg'] or 0
    
    if not request.user.is_administrador():
        total_analysts = User.objects.filter(role='analista', ativo=True, department=request.user.department).count()
    else:
        if selected_dept_id:
            total_analysts = User.objects.filter(role='analista', ativo=True, department_id=selected_dept_id).count()
        else:
            total_analysts = User.objects.filter(role='analista', ativo=True).count()
            
    complaints_without_analyst = queryset.filter(analista__isnull=True).count()
    
    # Reclamações urgentes (pendentes há mais de 3 dias)
    urgent_date = timezone.now().date() - timedelta(days=3)
    urgent_complaints = queryset.filter(
        status='pendente',
        data_reclamacao__lte=urgent_date
    ).count()
    
    context = {
        'total_complaints': total_complaints,
        'pending': pending,
        'em_replica': em_replica,
        'em_andamento': em_andamento,
        'resolved': resolved,
        'awaiting': awaiting,
        'recent_complaints': recent_complaints,
        'top_stores': list(top_stores),
        'top_stores_ranking': list(top_stores_ranking),
        'satisfaction_by_store': list(satisfaction_by_store),
        'complaints_by_status': list(complaints_by_status),
        'avg_satisfaction': round(avg_satisfaction, 1),
        'total_analysts': total_analysts,
        'complaints_without_analyst': complaints_without_analyst,
        'urgent_complaints': urgent_complaints,
    }
    return render(request, 'core/dashboard.html', context)


@login_required
def reports_view(request):
    """Página de relatórios e estatísticas avançadas"""
    from datetime import datetime, timedelta
    from django.db.models import Count, Avg, Q
    
    # Filtro base por departamento
    if request.user.is_administrador():
        base_queryset = Complaint.objects.all()
    else:
        base_queryset = Complaint.objects.filter(department=request.user.department)

    # Filtros de data
    date_from = request.GET.get('date_from', '')
    date_to = request.GET.get('date_to', '')
    
    # Base queryset
    complaints = base_queryset
    
    if date_from:
        try:
            date_from_obj = datetime.strptime(date_from, '%Y-%m-%d').date()
            complaints = complaints.filter(data_reclamacao__gte=date_from_obj)
        except:
            pass
    
    if date_to:
        try:
            date_to_obj = datetime.strptime(date_to, '%Y-%m-%d').date()
            complaints = complaints.filter(data_reclamacao__lte=date_to_obj)
        except:
            pass
    
    # Estatísticas gerais
    total = complaints.count()
    by_status = complaints.values('status').annotate(count=Count('id'))
    by_tipo = complaints.values('tipo_reclamacao').annotate(count=Count('id')).exclude(tipo_reclamacao__isnull=True)
    
    # Estatísticas por analista
    by_analyst = complaints.filter(analista__isnull=False).values(
        'analista__username', 'analista__first_name', 'analista__last_name'
    ).annotate(
        total=Count('id'),
        resolvidas=Count('id', filter=Q(status='resolvido')),
        pendentes=Count('id', filter=Q(status='pendente')),
        em_replica=Count('id', filter=Q(status='em_replica')),
        media_nota=Avg('nota_satisfacao')
    ).order_by('-total')
    
    # Estatísticas por loja
    by_store = complaints.values('loja_cod').annotate(
        total=Count('id'),
        resolvidas=Count('id', filter=Q(status='resolvido')),
        media_nota=Avg('nota_satisfacao')
    ).order_by('-total')[:20]
    
    # Satisfação do cliente
    satisfacao_stats = {
        'total_avaliacoes': complaints.filter(nota_satisfacao__isnull=False).count(),
        'media_geral': complaints.aggregate(avg=Avg('nota_satisfacao'))['avg'] or 0,
        'voltaria_sim': complaints.filter(volta_fazer_negocio='sim').count(),
        'voltaria_nao': complaints.filter(volta_fazer_negocio='nao').count(),
    }
    
    # Reclamações por período (últimos 30 dias)
    # Reclamações por período (últimos 30 dias) - Otimizado
    complaints_by_day = []
    days = 30
    date_threshold = timezone.now().date() - timedelta(days=days)
    
    daily_counts = complaints.filter(
        data_reclamacao__gte=date_threshold
    ).values('data_reclamacao').annotate(count=Count('id'))
    
    counts_map = {item['data_reclamacao']: item['count'] for item in daily_counts if item['data_reclamacao']}
    
    for i in range(days):
        date = timezone.now().date() - timedelta(days=days-1-i)
        count = counts_map.get(date, 0)
        complaints_by_day.append({'date': date.isoformat(), 'count': count})
    
    # Top problemas
    top_problemas = complaints.values('tipo_reclamacao').annotate(
        count=Count('id')
    ).exclude(tipo_reclamacao__isnull=True).order_by('-count')[:10]
    
    context = {
        'total': total,
        'by_status': by_status,
        'by_tipo': by_tipo,
        'by_analyst': by_analyst,
        'by_store': by_store,
        'satisfacao_stats': satisfacao_stats,
        'complaints_by_day': complaints_by_day,
        'top_problemas': top_problemas,
        'date_from': date_from,
        'date_to': date_to,
    }
    
    return render(request, 'core/reports.html', context)


@login_required
def complaint_list(request):
    # Filtro base por departamento
    selected_dept_id = request.session.get('selected_department_id')
    
    if request.user.is_administrador():
        if selected_dept_id:
            complaints = Complaint.objects.filter(department_id=selected_dept_id).select_related('analista', 'department')
        else:
            complaints = Complaint.objects.all().select_related('analista', 'department')
    else:
        complaints = Complaint.objects.filter(department=request.user.department).select_related('analista')
    
    # Filtros
    search = request.GET.get('search', '')
    status_filter = request.GET.get('status', '')
    loja_filter = request.GET.get('loja', '')
    
    if search:
        # Remover formatação do CPF se houver (pontos e traços)
        search_clean = re.sub(r'[^\d\w\s@.-]', '', search)  # Remove apenas pontos e traços, mantém números e letras
        
        # Se parecer um CPF (11 dígitos), buscar apenas números
        numbers_only = re.sub(r'\D', '', search_clean)
        if len(numbers_only) == 11:
            # Buscar CPF sem formatação
            complaints = complaints.filter(
                Q(id_ra__icontains=search) |
                Q(cpf_cliente__icontains=numbers_only) |
                Q(nome_cliente__icontains=search) |
                Q(email_cliente__icontains=search)
            )
        else:
            # Busca normal
            complaints = complaints.filter(
                Q(id_ra__icontains=search) |
                Q(cpf_cliente__icontains=search_clean) |
                Q(nome_cliente__icontains=search) |
                Q(email_cliente__icontains=search)
            )
    
    if status_filter:
        complaints = complaints.filter(status=status_filter)
    
    if loja_filter:
        complaints = complaints.filter(loja_cod=loja_filter)
    
    # Filtro de reclamações urgentes (pendentes há mais de 3 dias)
    urgentes = request.GET.get('urgentes', '')
    if urgentes == 'true':
        urgent_date = timezone.now().date() - timedelta(days=3)
        complaints = complaints.filter(
            status='pendente',
            data_reclamacao__lte=urgent_date
        )
    
    # Filtros avançados
    tipo_filter = request.GET.get('tipo', '')
    analista_filter = request.GET.get('analista', '')
    date_from = request.GET.get('date_from', '')
    date_to = request.GET.get('date_to', '')
    sem_analista = request.GET.get('sem_analista', '')
    
    if tipo_filter:
        complaints = complaints.filter(tipo_reclamacao=tipo_filter)
    
    if analista_filter:
        if analista_filter == 'sem_analista':
            complaints = complaints.filter(analista__isnull=True)
        else:
            complaints = complaints.filter(analista_id=analista_filter)
    
    if date_from:
        try:
            from datetime import datetime
            date_from_obj = datetime.strptime(date_from, '%Y-%m-%d').date()
            complaints = complaints.filter(data_reclamacao__gte=date_from_obj)
        except:
            pass
    
    if date_to:
        try:
            from datetime import datetime
            date_to_obj = datetime.strptime(date_to, '%Y-%m-%d').date()
            complaints = complaints.filter(data_reclamacao__lte=date_to_obj)
        except:
            pass
    
    if sem_analista == 'true':
        complaints = complaints.filter(analista__isnull=True)
    
    complaints = complaints.order_by('-created_at')
    
    # Contador de reclamações por responsável - analistas e gestores do depto
    if request.user.is_administrador():
        if selected_dept_id:
            base_analysts_stats = Complaint.objects.filter(department_id=selected_dept_id)
            analistas_list = User.objects.filter(role__in=['analista', 'gestor'], ativo=True, department_id=selected_dept_id).order_by('first_name')
        else:
            base_analysts_stats = Complaint.objects.all()
            analistas_list = User.objects.filter(role__in=['analista', 'gestor'], ativo=True).order_by('first_name')
    else:
        base_analysts_stats = Complaint.objects.filter(department=request.user.department)
        analistas_list = User.objects.filter(role__in=['analista', 'gestor'], ativo=True, department=request.user.department).order_by('first_name')

    complaints_by_analyst = base_analysts_stats.filter(
        analista__isnull=False
    ).values('analista__username', 'analista__first_name', 'analista__last_name').annotate(
        count=Count('id')
    ).order_by('-count')
    
    paginator = Paginator(complaints, 25)
    page = request.GET.get('page')
    complaints = paginator.get_page(page)
    
    context = {
        'complaints': complaints,
        'complaints_by_analyst': complaints_by_analyst,
        'analistas_list': analistas_list,
    }
    
    return render(request, 'core/complaint_list.html', context)



@login_required
def user_access_history(request):
    """Exibe o histórico de acessos dos usuários (apenas Administradores)"""
    if not request.user.is_administrador():
        messages.error(request, 'Acesso não autorizado.')
        return redirect('dashboard')
    
    # Logs de login
    login_logs = AuditLog.objects.filter(action='login').select_related('usuario').order_by('-created_at')
    
    # Estatísticas por usuário
    user_stats = User.objects.filter(ativo=True).annotate(
        total_logins=Count('audit_logs', filter=Q(audit_logs__action='login')),
        last_login_audit=Max('audit_logs__created_at', filter=Q(audit_logs__action='login'))
    ).order_by('-total_logins')
    
    # Paginação dos logs detalhados
    paginator = Paginator(login_logs, 50)
    page_number = request.GET.get('page')
    page_obj = paginator.get_page(page_number)
    
    context = {
        'page_obj': page_obj,
        'user_stats': user_stats,
    }
    
    return render(request, 'core/user_access_history.html', context)


@login_required
def store_list(request):
    # Filtro base por departamento
    selected_dept_id = request.session.get('selected_department_id')
    
    if request.user.is_administrador():
        if selected_dept_id:
            queryset = Complaint.objects.filter(department_id=selected_dept_id)
        else:
            queryset = Complaint.objects.all()
    else:
        queryset = Complaint.objects.filter(department=request.user.department)

    """Lista todas as lojas com reclamações, com filtros e ordenação"""
    # Obter todas as lojas com contagem de reclamações
    stores = queryset.values('loja_cod').annotate(
        count=Count('id'),
        pendentes=Count('id', filter=Q(status='pendente')),
        em_andamento=Count('id', filter=Q(status='em_andamento')),
        resolvidas=Count('id', filter=Q(status='resolvido')),
        aguardando=Count('id', filter=Q(status='aguardando_avaliacao'))
    )
    
    # Filtros
    search = request.GET.get('search', '')
    min_occurrences = request.GET.get('min_occurrences', '')
    
    if search:
        stores = stores.filter(loja_cod__icontains=search)
    
    if min_occurrences:
        try:
            min_count = int(min_occurrences)
            stores = stores.filter(count__gte=min_count)
        except ValueError:
            pass
    
    # Ordenação
    order_by = request.GET.get('order_by', 'count')
    order_direction = request.GET.get('order_direction', 'desc')
    
    if order_by == 'loja':
        if order_direction == 'asc':
            stores = stores.order_by('loja_cod')
        else:
            stores = stores.order_by('-loja_cod')
    elif order_by == 'count':
        if order_direction == 'asc':
            stores = stores.order_by('count')
        else:
            stores = stores.order_by('-count')
    else:
        stores = stores.order_by('-count')
    
    # Contar total antes da paginação
    total_stores = stores.count()
    
    # Paginação - aplicar limit e offset manualmente
    page = request.GET.get('page', 1)
    try:
        page = int(page)
    except (ValueError, TypeError):
        page = 1
    
    per_page = 20
    start = (page - 1) * per_page
    end = start + per_page
    
    stores_list = list(stores[start:end])
    
    # Calcular informações de paginação
    total_pages = (total_stores + per_page - 1) // per_page
    
    context = {
        'stores': stores_list,
        'search': search,
        'min_occurrences': min_occurrences,
        'order_by': order_by,
        'order_direction': order_direction,
        'total_stores': total_stores,
        'page': page,
        'total_pages': total_pages,
        'has_previous': page > 1,
        'has_next': page < total_pages,
        'previous_page': page - 1 if page > 1 else None,
        'next_page': page + 1 if page < total_pages else None,
    }
    
    return render(request, 'core/store_list.html', context)


@login_required
def store_complaints(request, loja_cod):
    """Lista todas as reclamações de uma loja específica"""
    # Filtro base por departamento
    selected_dept_id = request.session.get('selected_department_id')
    
    if request.user.is_administrador():
        if selected_dept_id:
            base_queryset = Complaint.objects.filter(department_id=selected_dept_id)
        else:
            base_queryset = Complaint.objects.all()
    else:
        base_queryset = Complaint.objects.filter(department=request.user.department)

    complaints = base_queryset.filter(loja_cod=loja_cod).select_related('analista').order_by('-created_at')
    
    # Estatísticas da loja
    store_stats = base_queryset.filter(loja_cod=loja_cod).aggregate(
        total=Count('id'),
        pendentes=Count('id', filter=Q(status='pendente')),
        em_andamento=Count('id', filter=Q(status='em_andamento')),
        resolvidas=Count('id', filter=Q(status='resolvido')),
        aguardando=Count('id', filter=Q(status='aguardando_avaliacao'))
    )
    
    # Filtros
    search = request.GET.get('search', '')
    status_filter = request.GET.get('status', '')
    
    if search:
        complaints = complaints.filter(
            Q(id_ra__icontains=search) |
            Q(cpf_cliente__icontains=search) |
            Q(nome_cliente__icontains=search) |
            Q(email_cliente__icontains=search)
        )
    
    if status_filter:
        complaints = complaints.filter(status=status_filter)
    
    # Paginação
    paginator = Paginator(complaints, 25)
    page = request.GET.get('page')
    complaints_page = paginator.get_page(page)
    
    context = {
        'loja_cod': loja_cod,
        'complaints': complaints_page,
        'store_stats': store_stats,
        'search': search,
        'status_filter': status_filter,
    }
    
    return render(request, 'core/store_complaints.html', context)


@login_required
def complaint_detail(request, pk):
    # Otimização: carregar departamento e analista em uma única query
    complaint = get_object_or_404(
        Complaint.objects.select_related('department', 'analista'), 
        pk=pk
    )
    
    # Trava de segurança por departamento
    if not request.user.is_administrador():
        if complaint.department != request.user.department:
            messages.error(request, 'Você não tem permissão para acessar esta reclamação.')
            return redirect('dashboard')
            
    # Atividades já estão com select_related('usuario')
    activities = complaint.activities.select_related('usuario').order_by('-created_at')

    # Adicionar comentário interno rápido
    if request.method == 'POST' and 'comentario_interno' in request.POST:
        comentario = request.POST.get('comentario_interno', '').strip()
        if comentario:
            Activity.objects.create(
                complaint=complaint,
                usuario=request.user,
                comentario=comentario,
                tipo_interacao='comentario_interno'
            )
            messages.success(request, 'Comentário adicionado com sucesso!')
            return redirect('complaint_detail', pk=pk)
    
    # Mudança rápida de status
    if request.method == 'POST' and 'novo_status' in request.POST:
        novo_status = request.POST.get('novo_status', '').strip()
        if novo_status and novo_status != complaint.status:
            status_antigo = complaint.get_status_display()
            complaint.status = novo_status
            complaint.save()
            Activity.objects.create(
                complaint=complaint,
                usuario=request.user,
                comentario=f'Status alterado de "{status_antigo}" para "{complaint.get_status_display()}"',
                tipo_interacao='mudanca_status'
            )
            messages.success(request, f'Status alterado para "{complaint.get_status_display()}"!')
            return redirect('complaint_detail', pk=pk)
    
    # Atribuição rápida de analista
    if request.method == 'POST' and 'novo_analista' in request.POST:
        # Atribuição deve respeitar o departamento
        novo_analista_id = request.POST.get('novo_analista', '').strip()
        if novo_analista_id:
            try:
                if request.user.is_administrador():
                    novo_analista = User.objects.get(id=novo_analista_id, role__in=['analista', 'gestor'], ativo=True, department=complaint.department)
                else:
                    novo_analista = User.objects.get(id=novo_analista_id, role__in=['analista', 'gestor'], ativo=True, department=request.user.department)
                    
                analista_antigo = complaint.analista.username if complaint.analista else "Não atribuído"
                complaint.analista = novo_analista
                complaint.save()
                Activity.objects.create(
                    complaint=complaint,
                    usuario=request.user,
                    comentario=f'Responsável alterado de "{analista_antigo}" para "{novo_analista.username}"',
                    tipo_interacao='atualizacao'
                )
                messages.success(request, f'Responsável atribuído: {novo_analista.username}!')
                return redirect('complaint_detail', pk=pk)
            except User.DoesNotExist:
                messages.error(request, 'Responsável não encontrado ou pertence a outro departamento!')
    
    # Lista de responsáveis para atribuição rápida (analistas e gestores do depto)
    # Otimização: carregar departamento para evitar query no template (se houver loop)
    if request.user.is_administrador():
        analistas_list = User.objects.filter(
            role__in=['analista', 'gestor'], 
            ativo=True, 
            department=complaint.department
        ).select_related('department').order_by('first_name')
    else:
        analistas_list = User.objects.filter(
            role__in=['analista', 'gestor'], 
            ativo=True, 
            department=request.user.department
        ).select_related('department').order_by('first_name')
    
    response = render(request, 'core/complaint_detail.html', {
        'complaint': complaint,
        'activities': activities,
        'analistas_list': analistas_list,
    })
    return response


@login_required
def complaint_create(request):
    # Trava de segurança: apenas Administradores ou o departamento 'CS Clientes' podem criar reclamações
    if not request.user.is_administrador():
        if not request.user.department or request.user.department.name != 'CS Clientes':
            messages.error(request, 'Apenas o departamento de CS Clientes e administradores podem criar reclamações.')
            # Redirecionar para a lista de reclamações ou dashboard
            return redirect('complaint_list')

    if request.method == 'POST':
        form = ComplaintForm(request.POST, user=request.user)
        if form.is_valid():
            complaint = form.save(commit=False)
            
            # Definir o departamento
            if request.user.is_administrador():
                selected_dept_id = request.session.get('selected_department_id')
                if selected_dept_id:
                    from .models import Department
                    complaint.department = Department.objects.filter(id=selected_dept_id).first()
                else:
                    # Se for admin e não tiver depto na sessão, buscar o padrão NRS
                    from .models import Department
                    complaint.department = Department.objects.filter(name='NRS Suporte').first()
            else:
                # Outros usuários usam seu departamento fixo
                complaint.department = request.user.department
            
            if request.user.is_analista():
                complaint.analista = request.user
            
            complaint.save()
            
            Activity.objects.create(
                complaint=complaint,
                usuario=request.user,
                comentario='Reclamação criada',
                tipo_interacao='criacao'
            )
            AuditLog.objects.create(
                usuario=request.user,
                action='create',
                target_type='Complaint',
                target_id=complaint.id
            )
            messages.success(request, 'Reclamação criada com sucesso!')
            return redirect('complaint_detail', pk=complaint.pk)
    else:
        form = ComplaintForm(user=request.user)
    return render(request, 'core/complaint_form.html', {'form': form})


@login_required
def complaint_edit(request, pk):
    complaint = get_object_or_404(Complaint, pk=pk)
    
    # Trava de segurança por departamento
    if not request.user.is_administrador():
        if complaint.department != request.user.department:
            messages.error(request, 'Você não tem permissão para editar esta reclamação.')
            return redirect('dashboard')
    if request.method == 'POST':
        form = ComplaintForm(request.POST, instance=complaint, user=request.user)
        if form.is_valid():
            # Capturar TODOS os valores antigos antes de salvar
            old_data = {
                'id_ra': complaint.id_ra,
                'cpf_cliente': complaint.cpf_cliente,
                'nome_cliente': complaint.nome_cliente,
                'sobrenome': complaint.sobrenome,
                'email_cliente': complaint.email_cliente,
                'telefone': complaint.telefone,
                'loja_cod': complaint.loja_cod,
                'origem_contato': complaint.origem_contato,
                'descricao': complaint.descricao,
                'status': complaint.status,
                'analista': complaint.analista,
                'data_reclamacao': complaint.data_reclamacao,
                'data_resposta': complaint.data_resposta,
                'nota_satisfacao': complaint.nota_satisfacao,
                'feedback_text': complaint.feedback_text,
            }
            
            # Salvar o formulário
            form.save()
            
            # Atualizar a instância do banco para comparar
            complaint.refresh_from_db()
            
            # Mapear nomes de campos para labels amigáveis
            field_labels = {
                'id_ra': 'ID RA',
                'cpf_cliente': 'CPF do Cliente',
                'nome_cliente': 'Nome do Cliente',
                'sobrenome': 'Sobrenome',
                'email_cliente': 'E-mail do Cliente',
                'telefone': 'Telefone',
                'loja_cod': 'Código da Loja',
                'origem_contato': 'Origem do Contato',
                'descricao': 'Descrição',
                'status': 'Status',
                'analista': 'Responsável',
                'data_reclamacao': 'Data da Reclamação',
                'data_resposta': 'Data de Resposta',
                'nota_satisfacao': 'Nota de Satisfação',
                'feedback_text': 'Feedback do Cliente',
            }
            
            # Criar atividades para cada campo alterado
            changes = []
            
            for field_name, old_value in old_data.items():
                new_value = getattr(complaint, field_name)
                
                # Comparar valores, tratando None, strings vazias e ForeignKey
                # Para ForeignKey, comparar os IDs
                if field_name == 'analista':
                    old_id = old_value.id if old_value else None
                    new_id = new_value.id if new_value else None
                    if old_id != new_id:
                        field_label = field_labels.get(field_name, field_name)
                        old_name = old_value.username if old_value else "Não atribuído"
                        new_name = new_value.username if new_value else "Não atribuído"
                        Activity.objects.create(
                            complaint=complaint,
                            usuario=request.user,
                            comentario=f'Campo "{field_label}" alterado de "{old_name}" para "{new_name}"',
                            tipo_interacao='atualizacao'
                        )
                        changes.append(field_name)
                    continue
                
                # Para outros campos, normalizar None e strings vazias
                old_val = old_value if old_value not in (None, "") else ""
                new_val = new_value if new_value not in (None, "") else ""
                
                if str(old_val).strip() != str(new_val).strip():
                    field_label = field_labels.get(field_name, field_name)
                    
                    # Formatação especial para alguns campos
                    if field_name == 'status':
                        old_display = dict(Complaint.STATUS_CHOICES).get(old_value, str(old_value))
                        new_display = complaint.get_status_display()
                        Activity.objects.create(
                            complaint=complaint,
                            usuario=request.user,
                            comentario=f'Campo "{field_label}" alterado de "{old_display}" para "{new_display}"',
                            tipo_interacao='mudanca_status'
                        )
                    elif field_name == 'origem_contato':
                        old_display = dict(Complaint.ORIGEM_CHOICES).get(old_value, str(old_value))
                        new_display = complaint.get_origem_contato_display()
                        Activity.objects.create(
                            complaint=complaint,
                            usuario=request.user,
                            comentario=f'Campo "{field_label}" alterado de "{old_display}" para "{new_display}"',
                            tipo_interacao='atualizacao'
                        )
                    elif field_name in ['data_reclamacao', 'data_resposta']:
                        old_str = old_value.strftime('%d/%m/%Y') if old_value else "Não informada"
                        new_str = new_value.strftime('%d/%m/%Y') if new_value else "Não informada"
                        Activity.objects.create(
                            complaint=complaint,
                            usuario=request.user,
                            comentario=f'Campo "{field_label}" alterado de "{old_str}" para "{new_str}"',
                            tipo_interacao='atualizacao'
                        )
                    else:
                        # Para outros campos, mostrar valores antigo e novo
                        old_str = str(old_value) if old_value else "(vazio)"
                        new_str = str(new_value) if new_value else "(vazio)"
                        Activity.objects.create(
                            complaint=complaint,
                            usuario=request.user,
                            comentario=f'Campo "{field_label}" alterado de "{old_str}" para "{new_str}"',
                            tipo_interacao='atualizacao'
                        )
                    changes.append(field_name)
            
            # Se não houver mudanças, criar atividade genérica
            if not changes:
                Activity.objects.create(
                    complaint=complaint,
                    usuario=request.user,
                    comentario='Informações da reclamação foram atualizadas',
                    tipo_interacao='atualizacao'
                )
            
            messages.success(request, 'Reclamação atualizada com sucesso!')
            return redirect('complaint_detail', pk=complaint.pk)
    else:
        form = ComplaintForm(instance=complaint, user=request.user)
    return render(request, 'core/complaint_form.html', {'form': form, 'complaint': complaint})


@login_required
def complaint_delete(request, pk):
    if not (request.user.is_gestor() or request.user.is_administrador()):
        messages.error(request, 'Você não tem permissão para excluir reclamações.')
        return redirect('complaint_list')
    
    complaint = get_object_or_404(Complaint, pk=pk)
    
    # Trava de segurança por departamento
    if not request.user.is_administrador():
        if complaint.department != request.user.department:
            messages.error(request, 'Você não tem permissão para excluir esta reclamação.')
            return redirect('dashboard')
    
    if request.method == 'POST':
        password = request.POST.get('password')
        
        if not password:
            messages.error(request, 'Por favor, informe sua senha para confirmar a exclusão.')
            return render(request, 'core/complaint_confirm_delete.html', {'complaint': complaint})
        
        # Verificar senha
        user = authenticate(request, username=request.user.username, password=password)
        if not user:
            messages.error(request, 'Senha incorreta. Tente novamente.')
            return render(request, 'core/complaint_confirm_delete.html', {'complaint': complaint})
        
        # Criar atividade antes de excluir
        Activity.objects.create(
            complaint=complaint,
            usuario=request.user,
            comentario=f'Reclamação {complaint.id_ra} foi excluída pelo gestor {request.user.username}',
            tipo_interacao='atualizacao'
        )
        
        AuditLog.objects.create(
            usuario=request.user,
            action='delete',
            target_type='Complaint',
            target_id=complaint.id,
            detalhes_json={'id_ra': complaint.id_ra}
        )
        complaint.delete()
        messages.success(request, 'Reclamação excluída com sucesso!')
        return redirect('complaint_list')
    
    return render(request, 'core/complaint_confirm_delete.html', {'complaint': complaint})


@login_required
def complaint_bulk_delete(request):
    """Exclusão em massa de reclamações - apenas para gestores"""
    if not (request.user.is_gestor() or request.user.is_administrador()):
        messages.error(request, 'Você não tem permissão para excluir reclamações.')
        return redirect('complaint_list')
    
    if request.method == 'POST':
        password = request.POST.get('password')
        delete_all = request.POST.get('delete_all') == 'on'
        selected_ids = request.POST.getlist('selected_complaints')
        
        if not password:
            messages.error(request, 'Por favor, informe sua senha para confirmar a exclusão.')
            return redirect('complaint_list')
        
        # Verificar senha
        user = authenticate(request, username=request.user.username, password=password)
        if not user:
            messages.error(request, 'Senha incorreta. Tente novamente.')
            return redirect('complaint_list')
        
        if delete_all:
            # Excluir todas as reclamações
            total = Complaint.objects.count()
            # Criar logs antes de excluir
            for complaint in Complaint.objects.all():
                AuditLog.objects.create(
                    usuario=request.user,
                    action='delete',
                    target_type='Complaint',
                    target_id=complaint.id,
                    detalhes_json={'id_ra': complaint.id_ra}
                )
            Complaint.objects.all().delete()
            messages.success(request, f'Todas as {total} reclamações foram excluídas com sucesso!')
        elif selected_ids:
            # Excluir selecionadas - pode vir como string separada por vírgula
            if isinstance(selected_ids, list) and len(selected_ids) > 0 and ',' in selected_ids[0]:
                selected_ids = selected_ids[0].split(',')
            # Converter para inteiros
            try:
                selected_ids = [int(id) for id in selected_ids if id]
            except (ValueError, TypeError):
                messages.error(request, 'IDs inválidos selecionados.')
                return redirect('complaint_list')
            
            complaints = Complaint.objects.filter(pk__in=selected_ids)
            count = complaints.count()
            for complaint in complaints:
                AuditLog.objects.create(
                    usuario=request.user,
                    action='delete',
                    target_type='Complaint',
                    target_id=complaint.id,
                    detalhes_json={'id_ra': complaint.id_ra}
                )
            complaints.delete()
            messages.success(request, f'{count} reclamação(ões) excluída(s) com sucesso!')
        else:
            messages.error(request, 'Nenhuma reclamação selecionada.')
        
        return redirect('complaint_list')
    
    return redirect('complaint_list')


def logout_view(request):
    """View customizada para logout que aceita GET"""
    logout(request)
    messages.success(request, 'Você saiu do sistema com sucesso!')
    return redirect('login')


@login_required
def user_list(request):
    """Lista de usuários - apenas para gestores e administradores"""
    if not (request.user.is_gestor() or request.user.is_administrador()):
        messages.error(request, 'Você não tem permissão para ver a lista de usuários.')
        return redirect('dashboard')
    
    from .models import Department
    departments = Department.objects.all().order_by('name')
    
    # Filtro base por departamento
    if request.user.is_administrador():
        users = User.objects.all().order_by('first_name', 'username')
    else:
        # Gestor só vê usuários do seu departamento
        users = User.objects.filter(department=request.user.department).order_by('first_name', 'username')
    
    search = request.GET.get('search', '')
    role_filter = request.GET.get('role', '')
    status_filter = request.GET.get('status', '')
    dept_filter = request.GET.get('department', '')
    
    if search:
        users = users.filter(
            Q(username__icontains=search) |
            Q(email__icontains=search) |
            Q(first_name__icontains=search) |
            Q(last_name__icontains=search)
        )
    
    if role_filter:
        users = users.filter(role=role_filter)
    
    if dept_filter:
        users = users.filter(department_id=dept_filter)
    
    if status_filter == 'ativo':
        users = users.filter(ativo=True)
    elif status_filter == 'inativo':
        users = users.filter(ativo=False)
    
    paginator = Paginator(users, 10)
    page = request.GET.get('page')
    users = paginator.get_page(page)
    
    context = {
        'users': users,
        'departments': departments
    }
    
    return render(request, 'core/user_list.html', context)


@login_required
@login_required
def user_create(request):
    """Criar novo usuário - apenas para gestores e administradores"""
    if not (request.user.is_gestor() or request.user.is_administrador()):
        messages.error(request, 'Você não tem permissão para criar usuários.')
        return redirect('dashboard')
    
    from .models import Department
    departments = Department.objects.all()
    
    if request.method == 'POST':
        username = request.POST.get('username')
        email = request.POST.get('email')
        password = request.POST.get('password')
        role = request.POST.get('role', 'analista')
        department_id = request.POST.get('department')
        ativo = request.POST.get('is_active') == 'on'
        first_name = request.POST.get('first_name', '')
        last_name = request.POST.get('last_name', '')
        profile_photo = request.FILES.get('profile_photo')
        
        # Restrições de Gestor
        if (request.user.is_gestor() or request.user.is_administrador()):
            role = 'analista' # Gestor só cria analista
            department_id = str(request.user.department_id) if request.user.department_id else None
            
        department = None
        if department_id:
            try:
                department = Department.objects.get(id=department_id)
            except (Department.DoesNotExist, ValueError):
                department = None
            
        if User.objects.filter(username=username).exists():
            messages.error(request, 'Este nome de usuário já existe.')
            return render(request, 'core/user_form.html', {
                'form_type': 'create',
                'departments': departments,
                'form_data': {
                    'email': email,
                    'first_name': first_name,
                    'last_name': last_name,
                    'role': role,
                    'ativo': ativo,
                    'department_id': department_id
                }
            })
        
        if User.objects.filter(email=email).exists():
            messages.error(request, 'Este e-mail já está cadastrado.')
            return render(request, 'core/user_form.html', {
                'form_type': 'create',
                'departments': departments,
                'form_data': {
                    'username': username,
                    'first_name': first_name,
                    'last_name': last_name,
                    'role': role,
                    'ativo': ativo,
                    'department_id': department_id
                }
            })
        
        user = User.objects.create_user(
            username=username,
            email=email,
            password=password,
            role=role,
            department=department,
            ativo=ativo,
            first_name=first_name,
            last_name=last_name
        )
        
        # Adicionar foto de perfil se fornecida
        if profile_photo:
            user.profile_photo = profile_photo
            user.save()

        # Integração Escala NRS Suporte
        if role == 'analista' and department and department.name == 'NRS Suporte':
            try:
                from .models import AnalistaEscala
                formatted_name = AnalistaEscala.format_schedule_name(first_name, last_name)
                AnalistaEscala.objects.create(
                    user=user,
                    nome=formatted_name,
                    ativo=ativo
                )
            except Exception as e:
                # Não impedir a criação do usuário se falhar a escala, mas logar
                print(f"Erro ao criar AnalistaEscala para {username}: {e}")
        
        AuditLog.objects.create(
            usuario=request.user,
            action='create',
            target_type='User',
            target_id=user.id,
            detalhes_json={'username': username, 'role': role}
        )
        
        messages.success(request, f'Usuário {username} criado com sucesso!')
        return redirect('user_list')
    
    return render(request, 'core/user_form.html', {
        'form_type': 'create',
        'departments': departments,
        'form_data': defaultdict(str)
    })


@login_required
def user_edit(request, pk):
    """Editar usuário - apenas para gestores e administradores"""
    if not (request.user.is_gestor() or request.user.is_administrador()):
        messages.error(request, 'Você não tem permissão para editar usuários.')
        return redirect('dashboard')
    
    from .models import Department
    departments = Department.objects.all()
    user_to_edit = get_object_or_404(User, pk=pk)
    
    # Gestor só edita usuários do seu depto e apenas analistas (ou a si mesmo)
    if request.user.is_gestor():
        if user_to_edit != request.user:
            if user_to_edit.department != request.user.department or user_to_edit.role != 'analista':
                messages.error(request, 'Você não tem permissão para editar este usuário.')
                return redirect('user_list')
    
    if request.method == 'POST':
        user_to_edit.email = request.POST.get('email')
        role = request.POST.get('role')
        department_id = request.POST.get('department')
        profile_photo = request.FILES.get('profile_photo')
        
        # Administrador pode mudar tudo, Gestor não muda role nem depto
        if request.user.is_administrador():
            user_to_edit.role = role
            if department_id:
                try:
                    user_to_edit.department = Department.objects.get(id=department_id)
                except (Department.DoesNotExist, ValueError):
                    user_to_edit.department = None
            else:
                user_to_edit.department = None
        
        user_to_edit.ativo = request.POST.get('is_active') == 'on'
        user_to_edit.first_name = request.POST.get('first_name', '')
        user_to_edit.last_name = request.POST.get('last_name', '')
        
        # Atualizar foto de perfil se fornecida
        if profile_photo:
            user_to_edit.profile_photo = profile_photo
        
        password = request.POST.get('password')
        if password:
            user_to_edit.set_password(password)
            
        user_to_edit.save()
        
        # Sincronização Escala NRS Suporte
        try:
            if hasattr(user_to_edit, 'escala_perfil'):
                analista_escala = user_to_edit.escala_perfil
                
                # Se saiu do NRS Suporte, desativa na escala
                if not user_to_edit.department or user_to_edit.department.name != 'NRS Suporte':
                    analista_escala.ativo = False
                else:
                    # Sincroniza dados
                    from .models import AnalistaEscala
                    analista_escala.nome = AnalistaEscala.format_schedule_name(user_to_edit.first_name, user_to_edit.last_name)
                    analista_escala.ativo = user_to_edit.ativo
                
                analista_escala.save()
        except Exception as e:
            print(f"Erro ao sincronizar AnalistaEscala para {user_to_edit.username}: {e}")
        
        AuditLog.objects.create(
            usuario=request.user,
            action='update',
            target_type='User',
            target_id=user_to_edit.id,
            detalhes_json={'username': user_to_edit.username, 'role': user_to_edit.role}
        )
        
        messages.success(request, f'Usuário {user_to_edit.username} atualizado!')
        return redirect('user_list')
    
    return render(request, 'core/user_form.html', {
        'form_type': 'edit',
        'target_user': user_to_edit,
        'departments': departments,
        'form_data': defaultdict(str)
    })


@login_required
def user_delete(request, pk):
    """Excluir usuário - apenas para gestores e administradores"""
    if not (request.user.is_gestor() or request.user.is_administrador()):
        messages.error(request, 'Você não tem permissão para excluir usuários.')
        return redirect('dashboard')
    
    try:
        user_to_delete = User.objects.get(pk=pk)
    except User.DoesNotExist:
        messages.error(request, 'Usuário não encontrado.')
        return redirect('user_list')
    
    # Restrições de Gestor
    if request.user.is_gestor():
        if user_to_delete.department != request.user.department or user_to_delete.role != 'analista':
            messages.error(request, 'Você não tem permissão para excluir este usuário.')
            return redirect('user_list')
            
    username = user_to_delete.username
    
    # Prevenir exclusão do próprio usuário logado
    if user_to_delete.id == request.user.id:
        messages.error(request, 'Você não pode excluir sua própria conta.')
        return redirect('user_list')
        
    user_to_delete.delete()
    
    AuditLog.objects.create(
        usuario=request.user,
        action='delete',
        target_type='User',
        target_id=pk,
        detalhes_json={'username': username}
    )
    
    messages.success(request, f'Usuário {username} excluído!')
    return redirect('user_list')

# Imports for settings_view
from django.contrib.auth.forms import PasswordChangeForm
from django.contrib.auth import update_session_auth_hash

@login_required
def settings_view(request):
    """Página de configurações do usuário"""
    if request.method == 'POST':
        if 'change_password' in request.POST:
            password_form = PasswordChangeForm(request.user, request.POST)
            if password_form.is_valid():
                user = password_form.save()
                update_session_auth_hash(request, user)  # Manter usuário logado
                messages.success(request, 'Sua senha foi alterada com sucesso!')
                return redirect('settings')
            else:
                messages.error(request, 'Erro ao alterar senha. Verifique os campos.')
    else:
        password_form = PasswordChangeForm(request.user)

    return render(request, 'core/settings.html', {
        'password_form': password_form
    })


@login_required
def export_complaints_csv(request):
    """Exportar reclamações para CSV"""
    # Filtro base por departamento
    selected_dept_id = request.session.get('selected_department_id')
    
    if request.user.is_administrador():
        if selected_dept_id:
            base_queryset = Complaint.objects.filter(department_id=selected_dept_id)
        else:
            base_queryset = Complaint.objects.all()
    else:
        base_queryset = Complaint.objects.filter(department=request.user.department)

    complaints = base_queryset.select_related('analista').all()
    
    # Aplicar filtros se existirem
    search = request.GET.get('search', '')
    status_filter = request.GET.get('status', '')
    loja_filter = request.GET.get('loja', '')
    
    if search:
        search_clean = re.sub(r'[^\d\w\s@.-]', '', search)
        numbers_only = re.sub(r'\D', '', search_clean)
        if len(numbers_only) == 11:
            complaints = complaints.filter(
                Q(id_ra__icontains=search) |
                Q(cpf_cliente__icontains=numbers_only) |
                Q(nome_cliente__icontains=search) |
                Q(email_cliente__icontains=search)
            )
        else:
            complaints = complaints.filter(
                Q(id_ra__icontains=search) |
                Q(cpf_cliente__icontains=search_clean) |
                Q(nome_cliente__icontains=search) |
                Q(email_cliente__icontains=search)
            )
    
    if status_filter:
        complaints = complaints.filter(status=status_filter)
    
    if loja_filter:
        complaints = complaints.filter(loja_cod=loja_filter)
    
    complaints = complaints.order_by('-created_at')
    
    response = HttpResponse(content_type='text/csv; charset=utf-8')
    response['Content-Disposition'] = f'attachment; filename="reclamacoes_{datetime.now().strftime("%Y%m%d_%H%M%S")}.csv"'
    
    writer = csv.writer(response)
    writer.writerow([
        'ID RA', 'CPF', 'Nome', 'Sobrenome', 'E-mail', 'Telefone',
        'Loja', 'Origem', 'Status', 'Analista', 'Data Reclamação',
        'Data Resposta', 'Nota Satisfação', 'Descrição'
    ])
    
    for complaint in complaints:
        writer.writerow([
            complaint.id_ra,
            complaint.cpf_cliente,
            complaint.nome_cliente,
            complaint.sobrenome or '',
            complaint.email_cliente,
            complaint.telefone or '',
            complaint.loja_cod,
            complaint.get_origem_contato_display(),
            complaint.get_status_display(),
            complaint.analista.username if complaint.analista else '',
            complaint.data_reclamacao.strftime('%d/%m/%Y') if complaint.data_reclamacao else '',
            complaint.data_resposta.strftime('%d/%m/%Y') if complaint.data_resposta else '',
            complaint.nota_satisfacao or '',
            complaint.descricao
        ])
    
    return response


@login_required
def export_complaints_xlsx(request):
    """Exportar reclamações para XLSX"""
    # Filtro base por departamento
    selected_dept_id = request.session.get('selected_department_id')
    
    if request.user.is_administrador():
        if selected_dept_id:
            base_queryset = Complaint.objects.filter(department_id=selected_dept_id)
        else:
            base_queryset = Complaint.objects.all()
    else:
        base_queryset = Complaint.objects.filter(department=request.user.department)

    complaints = base_queryset.select_related('analista').all()
    
    # Aplicar filtros se existirem
    search = request.GET.get('search', '')
    status_filter = request.GET.get('status', '')
    loja_filter = request.GET.get('loja', '')
    
    if search:
        search_clean = re.sub(r'[^\d\w\s@.-]', '', search)
        numbers_only = re.sub(r'\D', '', search_clean)
        if len(numbers_only) == 11:
            complaints = complaints.filter(
                Q(id_ra__icontains=search) |
                Q(cpf_cliente__icontains=numbers_only) |
                Q(nome_cliente__icontains=search) |
                Q(email_cliente__icontains=search)
            )
        else:
            complaints = complaints.filter(
                Q(id_ra__icontains=search) |
                Q(cpf_cliente__icontains=search_clean) |
                Q(nome_cliente__icontains=search) |
                Q(email_cliente__icontains=search)
            )
    
    if status_filter:
        complaints = complaints.filter(status=status_filter)
    
    if loja_filter:
        complaints = complaints.filter(loja_cod=loja_filter)
    
    complaints = complaints.order_by('-created_at')
    
    wb = Workbook()
    ws = wb.active
    ws.title = "Reclamações"
    
    # Cabeçalho
    headers = [
        'ID RA', 'CPF', 'Nome', 'Sobrenome', 'E-mail', 'Telefone',
        'Loja', 'Origem', 'Status', 'Analista', 'Data Reclamação',
        'Data Resposta', 'Nota Satisfação', 'Descrição'
    ]
    
    header_fill = PatternFill(start_color="2563EB", end_color="2563EB", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF", size=11)
    
    for col_num, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_num)
        cell.value = header
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal='center', vertical='center')
    
    # Dados
    for row_num, complaint in enumerate(complaints, 2):
        ws.cell(row=row_num, column=1, value=complaint.id_ra)
        ws.cell(row=row_num, column=2, value=complaint.cpf_cliente)
        ws.cell(row=row_num, column=3, value=complaint.nome_cliente)
        ws.cell(row=row_num, column=4, value=complaint.sobrenome or '')
        ws.cell(row=row_num, column=5, value=complaint.email_cliente)
        ws.cell(row=row_num, column=6, value=complaint.telefone or '')
        ws.cell(row=row_num, column=7, value=complaint.loja_cod)
        ws.cell(row=row_num, column=8, value=complaint.get_origem_contato_display())
        ws.cell(row=row_num, column=9, value=complaint.get_status_display())
        ws.cell(row=row_num, column=10, value=complaint.analista.username if complaint.analista else '')
        ws.cell(row=row_num, column=11, value=complaint.data_reclamacao.strftime('%d/%m/%Y') if complaint.data_reclamacao else '')
        ws.cell(row=row_num, column=12, value=complaint.data_resposta.strftime('%d/%m/%Y') if complaint.data_resposta else '')
        ws.cell(row=row_num, column=13, value=complaint.nota_satisfacao or '')
        ws.cell(row=row_num, column=14, value=complaint.descricao)
    
    # Ajustar largura das colunas
    column_widths = [12, 15, 20, 20, 25, 15, 10, 15, 15, 15, 15, 15, 10, 50]
    for col_num, width in enumerate(column_widths, 1):
        ws.column_dimensions[ws.cell(row=1, column=col_num).column_letter].width = width
    
    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = f'attachment; filename="reclamacoes_{datetime.now().strftime("%Y%m%d_%H%M%S")}.xlsx"'
    
    wb.save(response)
    return response


@login_required
def export_stores_csv(request):
    """Exportar lojas para CSV"""
    # Filtro base por departamento
    selected_dept_id = request.session.get('selected_department_id')
    
    if request.user.is_administrador():
        if selected_dept_id:
            base_queryset = Complaint.objects.filter(department_id=selected_dept_id)
        else:
            base_queryset = Complaint.objects.all()
    else:
        base_queryset = Complaint.objects.filter(department=request.user.department)

    stores = base_queryset.values('loja_cod').annotate(
        count=Count('id'),
        pendentes=Count('id', filter=Q(status='pendente')),
        em_andamento=Count('id', filter=Q(status='em_andamento')),
        resolvidas=Count('id', filter=Q(status='resolvido')),
        aguardando=Count('id', filter=Q(status='aguardando_avaliacao'))
    ).order_by('-count')
    
    response = HttpResponse(content_type='text/csv; charset=utf-8')
    response['Content-Disposition'] = f'attachment; filename="lojas_{datetime.now().strftime("%Y%m%d_%H%M%S")}.csv"'
    
    writer = csv.writer(response)
    writer.writerow(['Código da Loja', 'Total', 'Pendentes', 'Em Andamento', 'Aguardando Avaliação', 'Resolvidas'])
    
    for store in stores:
        writer.writerow([
            store['loja_cod'],
            store['count'],
            store['pendentes'],
            store['em_andamento'],
            store['aguardando'],
            store['resolvidas']
        ])
    
    return response


@login_required
def export_stores_xlsx(request):
    """Exportar lojas para XLSX"""
    # Filtro base por departamento
    selected_dept_id = request.session.get('selected_department_id')
    
    if request.user.is_administrador():
        if selected_dept_id:
            base_queryset = Complaint.objects.filter(department_id=selected_dept_id)
        else:
            base_queryset = Complaint.objects.all()
    else:
        base_queryset = Complaint.objects.filter(department=request.user.department)

    stores = base_queryset.values('loja_cod').annotate(
        count=Count('id'),
        pendentes=Count('id', filter=Q(status='pendente')),
        em_andamento=Count('id', filter=Q(status='em_andamento')),
        resolvidas=Count('id', filter=Q(status='resolvido')),
        aguardando=Count('id', filter=Q(status='aguardando_avaliacao'))
    ).order_by('-count')
    
    wb = Workbook()
    ws = wb.active
    ws.title = "Lojas"
    
    headers = ['Código da Loja', 'Total', 'Pendentes', 'Em Andamento', 'Aguardando Avaliação', 'Resolvidas']
    
    header_fill = PatternFill(start_color="2563EB", end_color="2563EB", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF", size=11)
    
    for col_num, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_num)
        cell.value = header
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal='center', vertical='center')
    
    for row_num, store in enumerate(stores, 2):
        ws.cell(row=row_num, column=1, value=store['loja_cod'])
        ws.cell(row=row_num, column=2, value=store['count'])
        ws.cell(row=row_num, column=3, value=store['pendentes'])
        ws.cell(row=row_num, column=4, value=store['em_andamento'])
        ws.cell(row=row_num, column=5, value=store['aguardando'])
        ws.cell(row=row_num, column=6, value=store['resolvidas'])
    
    column_widths = [20, 10, 12, 15, 20, 12]
    for col_num, width in enumerate(column_widths, 1):
        ws.column_dimensions[ws.cell(row=1, column=col_num).column_letter].width = width
    
    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = f'attachment; filename="lojas_{datetime.now().strftime("%Y%m%d_%H%M%S")}.xlsx"'
    
    wb.save(response)
    return response


@login_required
def export_users_csv(request):
    """Exportar usuários para CSV - apenas gestores"""
    if not (request.user.is_gestor() or request.user.is_administrador()):
        messages.error(request, 'Você não tem permissão para exportar usuários.')
        return redirect('dashboard')
    
    # Filtro por departamento
    if request.user.is_administrador():
        users = User.objects.all().order_by('username')
    else:
        users = User.objects.filter(department=request.user.department).order_by('username')
    
    response = HttpResponse(content_type='text/csv; charset=utf-8')
    response['Content-Disposition'] = f'attachment; filename="usuarios_{datetime.now().strftime("%Y%m%d_%H%M%S")}.csv"'
    
    writer = csv.writer(response)
    writer.writerow(['Username', 'E-mail', 'Nome', 'Sobrenome', 'Perfil', 'Ativo', 'Último Login', 'Data de Criação'])
    
    for user in users:
        writer.writerow([
            user.username,
            user.email,
            user.first_name or '',
            user.last_name or '',
            user.get_role_display(),
            'Sim' if user.ativo else 'Não',
            user.last_login.strftime('%d/%m/%Y %H:%M') if user.last_login else '',
            user.date_joined.strftime('%d/%m/%Y %H:%M') if user.date_joined else ''
        ])
    
    return response


@login_required
def export_users_xlsx(request):
    """Exportar usuários para XLSX - apenas gestores e administradores"""
    if not (request.user.is_gestor() or request.user.is_administrador()):
        messages.error(request, 'Você não tem permissão para exportar usuários.')
        return redirect('dashboard')
    
    # Filtro por departamento
    if request.user.is_administrador():
        users = User.objects.all().order_by('username')
    else:
        users = User.objects.filter(department=request.user.department).order_by('username')
    
    wb = Workbook()
    ws = wb.active
    ws.title = "Usuários"
    
    headers = ['Username', 'E-mail', 'Nome', 'Sobrenome', 'Perfil', 'Ativo', 'Último Login', 'Data de Criação']
    
    header_fill = PatternFill(start_color="2563EB", end_color="2563EB", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF", size=11)
    
    for col_num, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_num)
        cell.value = header
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal='center', vertical='center')
    
    for row_num, user in enumerate(users, 2):
        ws.cell(row=row_num, column=1, value=user.username)
        ws.cell(row=row_num, column=2, value=user.email)
        ws.cell(row=row_num, column=3, value=user.first_name or '')
        ws.cell(row=row_num, column=4, value=user.last_name or '')
        ws.cell(row=row_num, column=5, value=user.get_role_display())
        ws.cell(row=row_num, column=6, value='Sim' if user.ativo else 'Não')
        ws.cell(row=row_num, column=7, value=user.last_login.strftime('%d/%m/%Y %H:%M') if user.last_login else '')
        ws.cell(row=row_num, column=8, value=user.date_joined.strftime('%d/%m/%Y %H:%M') if user.date_joined else '')
    
    column_widths = [20, 30, 20, 20, 15, 10, 20, 20]
    for col_num, width in enumerate(column_widths, 1):
        ws.column_dimensions[ws.cell(row=1, column=col_num).column_letter].width = width
    
    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = f'attachment; filename="usuarios_{datetime.now().strftime("%Y%m%d_%H%M%S")}.xlsx"'
    
    wb.save(response)
    return response

@login_required
def import_complaints_xlsx(request):
    """Importar reclamações de arquivo XLSX"""
    if not (request.user.is_gestor() or request.user.is_administrador()):
        messages.error(request, 'Você não tem permissão para importar dados.')
        return redirect('dashboard')
    
    from .models import Department
    # Se for gestor, usa o depto dele. Se for admin, tenta pegar do POST ou padrão CS Clientes
    target_dept = request.user.department
    if request.user.is_administrador():
        dept_id = request.POST.get('department')
        if dept_id:
            target_dept = Department.objects.filter(id=dept_id).first()
        if not target_dept:
            target_dept = Department.objects.filter(slug='cs-clientes').first()

    if request.method == 'POST':
        if 'xlsx_file' not in request.FILES:
            messages.error(request, 'Por favor, selecione um arquivo XLSX.')
            return render(request, 'core/import_complaints.html')
        
        file = request.FILES['xlsx_file']
        if not file.name.endswith(('.xlsx', '.xls')):
            messages.error(request, 'Por favor, selecione um arquivo XLSX válido.')
            return render(request, 'core/import_complaints.html')
        
        try:
            from openpyxl import load_workbook
            wb = load_workbook(file, data_only=True)
            ws = wb.active
            
            imported = 0
            updated = 0
            skipped = 0
            errors = []
            total_rows = 0
            
            # Contar total de linhas (exceto cabeçalho)
            for row in ws.iter_rows(min_row=2, values_only=True):
                total_rows += 1
            
            # Começar da linha 2 (linha 1 é cabeçalho)
            for row_num, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
                status, msg = _process_complaint_row(row, row_num, target_dept)
                
                if status == 'created':
                    imported += 1
                elif status == 'updated':
                    updated += 1
                elif status == 'skipped':
                    skipped += 1
                    errors.append(msg)
                elif status == 'error':
                    errors.append(msg)
            
            # Mensagens
            total_processed = imported + updated
            if total_processed > 0:
                msg = f"Importação concluída! Processadas {total_rows} linha(s) da planilha. "
                msg += f"{imported} reclamação(ões) criada(s), {updated} atualizada(s)."
                if skipped > 0:
                    msg += f" {skipped} linha(s) ignorada(s) (ID RA vazio)."
                if len(errors) > skipped:
                    msg += f" {len(errors) - skipped} aviso(s)."
                messages.success(request, msg)
            else:
                messages.error(request, f"Nenhuma reclamação foi importada. Verifique os erros abaixo.")
            
            if errors:
                for error in errors[:15]:
                    messages.warning(request, error)
                if len(errors) > 15:
                    messages.warning(request, f"... e mais {len(errors) - 15} aviso(s)/erro(s).")
            
            return redirect('complaint_list')
            
        except Exception as e:
            messages.error(request, f'Erro ao processar arquivo: {str(e)}')
            return render(request, 'core/import_complaints.html')
    
    return render(request, 'core/import_complaints.html')


def _process_complaint_row(row, row_num, target_dept):
    """
    Processa uma linha da planilha de importação.
    Retorna (status, mensagem)
    status: 'created', 'updated', 'skipped', 'error'
    """
    try:
        # Mapeamento de status e tipos
        status_map = {
            'pendente': 'pendente', 'em andamento': 'em_andamento', 'em réplica': 'em_replica',
            'aguardando avaliação': 'aguardando_avaliacao', 'resolvido': 'resolvido', 'resolvida': 'resolvido',
        }
        tipo_map = {
            'nota fiscal': 'nota_fiscal', 'pagamento não processado - cartão': 'pagamento_cartao',
            'pagamento não processado - pix': 'pagamento_pix', 'pagamento não processado - checkout web': 'pagamento_checkout',
            'assinatura mensal': 'assinatura_mensal', 'lavagem': 'lavagem', 'secagem': 'secagem',
            'atendimento': 'atendimento', 'sistema/totem': 'sistema_totem', 'totem': 'sistema_totem',
            'cupons': 'cupons', 'outros': 'outros',
        }
        volta_negocio_map = {'sim': 'sim', 's': 'sim', 'não': 'nao', 'nao': 'nao', 'n': 'nao'}

        # Mapear colunas
        loja_cod = str(row[0]).strip() if len(row) > 0 and row[0] else 'Não informado'
        nome_completo = str(row[1]).strip() if len(row) > 1 and row[1] else 'Nome não informado'
        id_ra = str(row[2]).strip() if len(row) > 2 and row[2] else None
        
        # Validação ID RA
        if not id_ra or id_ra == '':
            return 'skipped', f"Linha {row_num}: ID RA está vazio - linha ignorada"

        cpf = str(row[3]).strip() if len(row) > 3 and row[3] else None
        email_cliente = str(row[4]).strip() if len(row) > 4 and row[4] else None
        telefone = str(row[5]).strip() if len(row) > 5 and row[5] else ''
        data_reclamacao = row[6] if len(row) > 6 and row[6] else None
        problema = str(row[7]).strip().lower() if len(row) > 7 and row[7] else None
        status = str(row[8]).strip().lower() if len(row) > 8 and row[8] else 'pendente'
        analista_nome = str(row[9]).strip() if len(row) > 9 and row[9] else None
        nota = row[10] if len(row) > 10 and row[10] else None
        volta_negocio = str(row[11]).strip().lower() if len(row) > 11 and row[11] else None

        # Processar CPF
        cpf_clean = re.sub(r'\D', '', str(cpf)) if cpf else '00000000000'
        if len(cpf_clean) != 11: cpf_clean = '00000000000'

        # Dividir nome
        nome_parts = str(nome_completo).split(maxsplit=1)
        nome_cliente = nome_parts[0] if nome_parts else 'Nome não informado'
        sobrenome = nome_parts[1] if len(nome_parts) > 1 else ''

        # Processar Data
        data_reclamacao_value = timezone.now().date()
        if data_reclamacao:
            if isinstance(data_reclamacao, datetime):
                data_reclamacao_value = data_reclamacao.date()
            elif isinstance(data_reclamacao, str) and data_reclamacao.strip():
                for fmt in ('%d/%m/%Y', '%Y-%m-%d', '%d-%m-%Y'):
                    try:
                        data_reclamacao_value = datetime.strptime(data_reclamacao.strip(), fmt).date()
                        break
                    except ValueError: pass

        # Mapear Status e Tipo
        status_value = status_map.get(status, 'pendente')
        tipo_reclamacao_value = tipo_map.get(problema, 'outros') if problema else None
        volta_negocio_value = volta_negocio_map.get(volta_negocio, 'nao_informado') if volta_negocio else None

        # Buscar Analista
        analista_obj = None
        if analista_nome and analista_nome.lower().strip() not in ['selecione um analista (opcional)', 'não atribuido', '', 'nao atribuido']:
            try:
                parts = str(analista_nome).strip().split()
                q = Q()
                for p in parts:
                    q |= Q(first_name__icontains=p) | Q(last_name__icontains=p)
                analista_obj = User.objects.filter(role='analista', ativo=True).filter(q).first()
            except: pass

        # Nota
        nota_value = None
        if nota is not None:
            try:
                nota_value = max(0, min(10, int(float(nota))))
            except: pass

        # Email fallback
        if not email_cliente or '@' not in email_cliente:
            email_cliente = f'{cpf_clean}@importado.com'

        descricao = f'Importado da planilha' + (f' - Tipo: {problema}' if problema else ' - Tipo: Não informado')

        complaint, created = Complaint.objects.update_or_create(
            id_ra=str(id_ra).strip(),
            defaults={
                'cpf_cliente': cpf_clean,
                'nome_cliente': nome_cliente,
                'sobrenome': sobrenome,
                'email_cliente': email_cliente,
                'telefone': telefone,
                'loja_cod': loja_cod,
                'origem_contato': 'RA',
                'descricao': descricao,
                'status': status_value,
                'analista': analista_obj,
                'data_reclamacao': data_reclamacao_value,
                'tipo_reclamacao': tipo_reclamacao_value,
                'nota_satisfacao': nota_value,
                'volta_fazer_negocio': volta_negocio_value,
                'department': target_dept,
            }
        )
        return ('created' if created else 'updated'), None

    except Exception as e:
        return 'error', f"Linha {row_num}: {str(e)}"


@login_required
def import_complaints_batch(request):
    """API para importação em lote (Batch)"""
    if request.method != 'POST':
        return JsonResponse({'error': 'Método não permitido'}, status=405)
        
    if not (request.user.is_gestor() or request.user.is_administrador()):
        return JsonResponse({'error': 'Permissão negada'}, status=403)

    try:
        import json
        data = json.loads(request.body)
        rows = data.get('rows', [])
        
        # Determinar departamento (mesma lógica do import normal)
        from .models import Department
        target_dept = request.user.department
        if request.user.is_administrador():
            # Tentar pegar dept do body se enviado, ou usar padrão
             target_dept = Department.objects.filter(slug='cs-clientes').first()

        results = {
            'created': 0,
            'updated': 0,
            'skipped': 0,
            'errors': []
        }

        for i, row in enumerate(rows):
            # row deve ser uma lista/array vinda do SheetJS header:1
            # O índice da linha real deve considerar offset do lote + cabeçalho se houver
            # Aqui usamos apenas index relativo ao batch para log
            db_status, msg = _process_complaint_row(row, i, target_dept)
            
            if db_status == 'created': results['created'] += 1
            elif db_status == 'updated': results['updated'] += 1
            elif db_status == 'skipped': results['skipped'] += 1
            elif db_status == 'error': results['errors'].append(msg)

        return JsonResponse(results)

    except Exception as e:
        return JsonResponse({'error': str(e)}, status=500)


@login_required
def store_audit_form(request, store_id):
    store = get_object_or_404(Store, id=store_id)
    
    if request.method == 'POST':
        # Processar o formulário de auditoria
        # ... (lógica de salvamento da auditoria)
        
        has_irregularity = False # Exemplo, você precisará determinar isso com base nos dados do formulário
        
        if has_irregularity:
            messages.warning(request, f"Auditoria da loja {store.code} finalizada com irregularidades detectadas. Gestor notificado.")
            request.session['play_irregularity_sound'] = True # Sinal para o front tocar som
        else:
            messages.success(request, f"Auditoria da loja {store.code} finalizada com sucesso (Tudo conforme).")
            
        return redirect('verificacao_lojas')

    # Checklist items
    items_choices = [
        ('cameras', 'Câmeras'),
        ('estofados', 'Estofados'),
        ('cestos_medidas', 'Cestos de medidas'),
        ('layout', 'Layout'),
        ('tv', 'TV'),
        ('totem', 'Totem'),
        ('limpeza', 'Limpeza da loja'),
        ('marketing', 'Marketing'),
    ]

    # Histórico de auditorias da loja
    history = StoreAudit.objects.filter(store=store).order_by('-created_at')[:10]
    
    context = {
        'store': store,
        'history': history,
        'items_choices': items_choices,
        'title': f'Auditoria: Loja {store.code}'
    }
    return render(request, 'core/store_audit_form.html', context)
    
@login_required
def under_development(request, page_name='Página'):
    return render(request, 'core/under_development.html', {'page_name': page_name})


@login_required
def sites_view(request):
    """Página de Sites e Sistemas - NRS Suporte"""
    if not request.user.is_administrador():
        if not request.user.department or request.user.department.name != 'NRS Suporte':
            messages.error(request, 'Você não tem permissão para acessar as ferramentas de NRS Suporte.')
            return redirect('dashboard')
            
    return render(request, 'core/sites.html')


@login_required
def localizacao_view(request):
    """Página de Localização das Lojas - NRS Suporte"""
    if not request.user.is_administrador():
        if not request.user.department or request.user.department.name != 'NRS Suporte':
            messages.error(request, 'Você não tem permissão para acessar as ferramentas de NRS Suporte.')
            return redirect('dashboard')
            
    return render(request, 'core/localizacao.html')


@login_required
def escala_view(request):
    """Página de Escala - NRS Suporte"""
    if not request.user.is_administrador():
        if not request.user.department or request.user.department.name != 'NRS Suporte':
            messages.error(request, 'Você não tem permissão para acessar as ferramentas de NRS Suporte.')
            return redirect('dashboard')
            
    from .models import Turno, AnalistaEscala, FolgaManual
    
    # --- Sincronização de Usuários (Comentada para performance, rodar via management command se necessário) ---
    # try:
    #     users_missing_profile = User.objects.filter(
    #         department__name='NRS Suporte', 
    #         role='analista', 
    #         ativo=True,
    #         escala_perfil__isnull=True
    #     )
    #     for user in users_missing_profile:
    #         formatted_name = AnalistaEscala.format_schedule_name(user.first_name, user.last_name)
    #         existing = AnalistaEscala.objects.filter(nome=formatted_name, user__isnull=True).first()
    #         if existing:
    #             existing.user = user
    #             existing.save()
    #         else:
    #             AnalistaEscala.objects.create(user=user, nome=formatted_name, ativo=True)
    # except Exception as e:
    #     print(f"Erro ao sincronizar analistas na escala: {e}")
    # ---------------------------------------------------

    turnos = Turno.objects.filter(ativo=True).order_by('ordem', 'nome')
    analistas = AnalistaEscala.objects.filter(ativo=True).select_related('turno').order_by('turno__ordem', 'ordem', 'nome')
    
    # Preparar dados para JSON
    turnos_data = [{
        'id': t.id,
        'nome': t.nome,
        'horario': t.horario,
        'cor': t.cor,
        'ordem': t.ordem
    } for t in turnos]
    
    analistas_data = [{
        'id': a.id,
        'nome': a.nome,
        'turno': a.turno.nome if a.turno else None,
        'turno_id': a.turno.id if a.turno else None,
        'pausa': a.pausa,
        'data_primeira_folga': a.data_primeira_folga.isoformat() if a.data_primeira_folga else None,
        'ordem': a.ordem
    } for a in analistas]
    
    # Folgas manuais como dicionário
    folgas = FolgaManual.objects.select_related('analista').all()
    folgas_data = {}
    for f in folgas:
        key = f"{f.analista.id}-{f.data.year}-{f.data.month}-{f.data.day}"
        folgas_data[key] = {
            'id': f.id,
            'tipo': f.tipo,
            'motivo': f.motivo
        }
    
    import json
    is_admin = request.user.is_gestor() or request.user.is_administrador()
    context = {
        'turnos_json': json.dumps(turnos_data),
        'analistas_json': json.dumps(analistas_data),
        'folgas_json': json.dumps(folgas_data),
        'is_admin': is_admin,
        'is_admin_json': 'true' if is_admin else 'false'
    }
    
    return render(request, 'core/escala.html', context)

@login_required
def calendar_view(request):
    """Visualização do calendário"""
    user = request.user
    # Gestores/Admins podem editar tudo
    is_manager = user.role in ['gestor', 'administrador']
    # Analistas do NRS podem criar/editar seus próprios eventos
    is_nrs_analyst = user.role == 'analista' and user.department and user.department.name == 'NRS Suporte'
    
    can_create = is_manager or is_nrs_analyst
    
    context = {
        'can_edit': can_create,  # No template, isso controla o botão "Adicionar"
        'user_id': user.id,
        'user_full_name': f"{user.first_name} {user.last_name}".strip() or user.username,
        'is_manager': is_manager,
    }
    return render(request, 'core/calendar.html', context)

@login_required
def knowledge_base_view(request):
    """Visualização da Base de Conhecimento"""
    can_edit = request.user.is_administrador() or request.user.is_gestor()
    
    context = {
        'can_edit': can_edit,
    }
    return render(request, 'core/knowledge_base.html', context)


@login_required
def performance_view(request):
    """Página de Desempenho do Time"""
    from .models import Department, IndicadorDesempenho
    import json
    
    user = request.user
    
    # Obter departamento NRS Suporte
    try:
        nrs_dept = Department.objects.get(name='NRS Suporte')
    except Department.DoesNotExist:
        nrs_dept = None
    
    # Determinar se é analista ou gestor/admin
    is_analista = user.role == 'analista'
    can_edit = user.role in ['gestor', 'administrador']
    
    # Título da página baseado no role
    if is_analista:
        page_title = "Meu Desempenho"
    else:
        page_title = "Desempenho do Time"
    
    # Lista de analistas (para dropdown do gestor)
    if can_edit and nrs_dept:
        analistas = User.objects.filter(
            department=nrs_dept
        ).order_by('first_name', 'username')
    else:
        analistas = []
    
    # Analista selecionado (para visualização)
    selected_analista_id = request.GET.get('analista_id')
    if is_analista:
        selected_analista_id = user.id
    elif not selected_analista_id and analistas:
        selected_analista_id = analistas.first().id if analistas.exists() else None
    
    if selected_analista_id:
        try:
            selected_analista_id = int(selected_analista_id)
        except (ValueError, TypeError):
            selected_analista_id = None

    # Filtro de período
    period = request.GET.get('period', '6')
    try:
        period_int = int(period) if period != 'all' else None
    except ValueError:
        period_int = 6

    # Obter KPIs do analista selecionado
    kpis_data = []
    if selected_analista_id and nrs_dept:
        kpis_query = IndicadorDesempenho.objects.filter(
            analista_id=selected_analista_id,
            department=nrs_dept
        ).order_by('ano', 'mes')
        
        # Se houver limite de período, pegar os últimos N meses
        if period_int:
            # Pegamos todos e depois fatiamos os últimos period_int para garantir que pegamos os mais recentes cronologicamente
            kpis = list(kpis_query)
            if len(kpis) > period_int:
                kpis = kpis[-period_int:]
        else:
            kpis = list(kpis_query)
        
        # Obter metas globais para o departamento
        metas_globais = {
            f"{m.mes:02d}/{m.ano}": m 
            for m in MetaMensalGlobal.objects.filter(department=nrs_dept)
        }
        
        for kpi in kpis:
            label = f"{kpi.mes:02d}/{kpi.ano}"
            meta = metas_globais.get(label)
            
            kpis_data.append({
                'id': kpi.id,
                'mes': kpi.mes,
                'ano': kpi.ano,
                'label': label,
                'nps': float(kpi.nps) if kpi.nps else None,
                'tme': kpi.tme,
                'chats': kpi.chats,
                # Metas individuais (caso queira manter como fallback)
                'meta_tme_ind': kpi.meta_tme,
                'meta_nps_ind': float(kpi.meta_nps) if kpi.meta_nps else None,
                'meta_chats_ind': kpi.meta_chats,
                # Metas globais (as que realmente importam agora)
                'meta_tme': meta.meta_tme if meta else None,
                'meta_nps': float(meta.meta_nps) if meta and meta.meta_nps else None,
                'meta_chats': meta.meta_chats if meta else None,
            })
    
    selected_analista = None
    if selected_analista_id:
        try:
            selected_analista = User.objects.get(id=selected_analista_id)
        except User.DoesNotExist:
            pass


    # Lista de analistas formatada com flag de seleção
    analistas_list = []
    for a in analistas:
        analistas_list.append({
            'id': a.id,
            'name': a.get_full_name() or a.username,
            'is_selected': a.id == selected_analista_id
        })

    context = {
        'page_title': page_title,
        'is_analista': is_analista,
        'can_edit': can_edit,
        'can_edit_str': "true" if can_edit else "false",
        'analistas_list': analistas_list,
        'selected_analista_id': selected_analista_id,
        'selected_analista': selected_analista,
        'kpis_json': json.dumps(kpis_data),
        'current_period': period,
        'analistas_json': json.dumps([{
            'id': a.id,
            'nome': a.get_full_name() or a.username
        } for a in analistas]),
    }
    
    return render(request, 'core/desempenho.html', context)


@login_required
def quadro_view(request):
    """Visualização do Quadro Kanban"""
    from .models import KanbanBoard, KanbanList, CardLabel, Department
    import json
    
    # Get or create board for user
    board = KanbanBoard.objects.filter(owner=request.user).first()
    
    if not board:
        # Try to find any board or create default
        board = KanbanBoard.objects.first()
        if not board:
            board = KanbanBoard.objects.create(
                name='Quadro Principal',
                owner=request.user,
                background_color='#2563eb'
            )
    
    # Create default lists if they don't exist
    if not board.lists.exists():
        KanbanList.objects.create(board=board, name='A Fazer', position=0)
        KanbanList.objects.create(board=board, name='Em Andamento', position=1)
        KanbanList.objects.create(board=board, name='Concluído', position=2)
    
    # Create default labels if they don't exist
    if not board.labels.exists():
        default_labels = [
            ('Urgente', '#ef4444'),
            ('Importante', '#f97316'),
            ('Normal', '#6b7280'),
            ('Baixa', '#3b82f6'),
            ('Reunião', '#6366f1'),
            ('Documentação', '#0891b2'),
            ('Bug', '#dc2626'),
            ('Feature', '#059669'),
            ('Concluído', '#22c55e'),
        ]
        for name, color in default_labels:
            CardLabel.objects.create(board=board, name=name, color=color)
    
    # Get lists and prefetch cards with labels
    listas = board.lists.filter(is_archived=False).prefetch_related('cards__labels').order_by('position')
    labels = board.labels.all()
    
    labels_data = [{'id': l.id, 'name': l.name, 'color': l.color} for l in labels]
    
    # Get NRS Suporte department users for members assignment
    nrs_dept = Department.objects.filter(name='NRS Suporte').first()
    members = []
    if nrs_dept:
        nrs_users = User.objects.filter(department=nrs_dept, ativo=True).order_by('first_name', 'username')
        members = [{
            'id': u.id,
            'name': u.get_full_name() or u.username,
            'initials': (u.first_name[:1] + u.last_name[:1]).upper() if u.first_name and u.last_name else u.username[:2].upper(),
            'role': u.get_role_display()
        } for u in nrs_users]
    
    context = {
        'board': board,
        'listas': listas,
        'labels_json': json.dumps(labels_data),
        'members_json': json.dumps(members),
    }
    return render(request, 'core/quadro.html', context)


@login_required
def tasks_view(request):
    """View para a aba de tarefas e rotina"""
    user = request.user
    is_manager = user.role in ['gestor', 'administrador']
    
    # Permissão para botão "Criar": Gestores/Admins OU Analistas do NRS Suporte
    is_nrs_analyst = (user.role == 'analista' and user.department and 
                      user.department.name == 'NRS Suporte')
    
    show_create_button = is_manager or is_nrs_analyst or user.is_administrador()
    
    context = {
        'title': 'Tarefas e Rotinas',
        'is_manager': is_manager,
        'show_create_button': show_create_button,
    }
    return render(request, 'core/tarefas.html', context)


@login_required
def solicitacoes_view(request):
    """View para a aba de solicitações de estorno (CS Clientes)"""
    user = request.user
    is_manager = user.role in ['gestor', 'administrador']
    is_cs_clientes = user.department and user.department.name == 'CS Clientes'
    
    context = {
        'title': 'Solicitações de Estorno',
        'is_manager': is_manager,
        'is_cs_clientes': is_cs_clientes,
    }
    return render(request, 'core/solicitacoes.html', context)


# ========================================
# VIEWS PARA VERIFICAÇÃO DE LOJAS (AUDITORIA)
# ========================================

@login_required
@ensure_csrf_cookie
def verificacao_lojas(request):
    """Página principal de verificação de lojas (NRS Suporte)"""
    from django.core.paginator import Paginator
    
    # 1. Base QuerySet
    # Tab handling
    search_query = request.GET.get('q')
    tab = request.GET.get('tab', 'all')  # DEFAULT: 'all' (Lojas Ativas)
    
    # If searching, search EVERYTHING (Active + Inactive) and force List View
    if search_query:
        stores_queryset = Store.objects.all().order_by('code')
        tab = 'all' # Force 'all' tab to ensure list view is rendered
    # If NOT searching, respect tab filters (Active vs Suspended)
    elif tab == 'suspended':
        stores_queryset = Store.objects.filter(active=False).order_by('code')
    else:
        stores_queryset = Store.objects.filter(active=True).order_by('code')
    
    # 2. Status Calculation (Aggregate counts efficiently)
    # OPTIMIZED: Use distinct() to avoid loading duplicate store_ids
    # This prevents WORKER TIMEOUT when there are thousands of audits
    verified_store_ids = set(StoreAudit.objects.values_list('store_id', flat=True).distinct())
    irregular_store_ids = set(StoreAuditIssue.objects.filter(status='aberta').values_list('store_id', flat=True).distinct())
    suspended_count = Store.objects.filter(active=False).count()
    
    # Calculate OK stores
    ok_store_ids = verified_store_ids - irregular_store_ids
    
    # 3. Apply Filters
    # Scope Filter (My Stores vs All Stores)
    # Default to 'all' for everyone to allow analysts to see all stores
    scope = request.GET.get('scope', 'all')
    if scope == 'my_stores' and request.user.is_authenticated:
        # Import dynamically to avoid circular dependency
        from .models import AnalystAssignment 
        my_store_ids = AnalystAssignment.objects.filter(
            analyst=request.user, 
            active=True
        ).values_list('store_id', flat=True)
        stores_queryset = stores_queryset.filter(id__in=my_store_ids)

    # Filter by Tab/Status
    if tab == 'irregular':
        stores_queryset = stores_queryset.filter(id__in=irregular_store_ids)
    elif tab == 'verified':
        stores_queryset = stores_queryset.filter(id__in=ok_store_ids)
    elif tab == 'suspended':
        # Already filtered by active=False above
        pass
    
    # Legacy 'filter' param support (optional, but good for backward compat link)
    filter_type = request.GET.get('filter')
    if filter_type == 'verified':
        stores_queryset = stores_queryset.filter(id__in=verified_store_ids)
    elif filter_type == 'ok':
        stores_queryset = stores_queryset.filter(id__in=ok_store_ids)
    elif filter_type == 'irregular':
        stores_queryset = stores_queryset.filter(id__in=irregular_store_ids)
    
    # Search Filter
    search_query = request.GET.get('q')
    if search_query:
        stores_queryset = stores_queryset.filter(
            Q(code__icontains=search_query) | 
            Q(city__icontains=search_query)
        )

    # Annonation removed - calculated in loop below for safety

    # 4. Pagination
    paginator = Paginator(stores_queryset, 25)  # 25 items per page
    page_number = request.GET.get('page')
    page_obj = paginator.get_page(page_number)
    
    # 5. Eager Load Latest Audits for CURRENT PAGE only (Performance Fix)
    page_store_ids = [store.id for store in page_obj]
    
    # Calculate Monthly Counts (Explicitly)
    from django.utils import timezone
    from django.db.models import Count
    now = timezone.now()
    start_of_month = now.replace(day=1, hour=0, minute=0, second=0, microsecond=0)
    
    monthly_counts = StoreAudit.objects.filter(
        store_id__in=page_store_ids,
        created_at__gte=start_of_month
    ).values('store_id').annotate(count=Count('id'))
    
    monthly_counts_map = {item['store_id']: item['count'] for item in monthly_counts}
    
    # Fetch latest audits (optimized)
    # Using window function or grouping in Python is needed for 'top N per group'
    # Since page size is small (25), we can fetch recent audits for these stores efficiently
    audits_qs = StoreAudit.objects.filter(store_id__in=page_store_ids).select_related('analyst').prefetch_related('items').order_by('store', '-created_at')
    
    # Group by store in memory
    from collections import defaultdict
    store_audits_map = defaultdict(list)
    for audit in audits_qs:
        # We only need the top 3 for the list view, others are loaded via API
        if len(store_audits_map[audit.store_id]) < 3:
            store_audits_map[audit.store_id].append(audit)
            
    latest_audits = {
        store_id: audits[0] if audits else None
        for store_id, audits in store_audits_map.items()
    }
    
    # Attach stats manually to avoid template queries
    for store in page_obj:
        store.latest_audit = latest_audits.get(store.id)
        store.audits_this_month_count = monthly_counts_map.get(store.id, 0)
        
        
        # Determine status for UI badge
        if not store.active:
            store.ui_status = 'suspended'
        elif store.id in irregular_store_ids:
            store.ui_status = 'irregular'
        elif store.id in verified_store_ids:
            store.ui_status = 'compliant'
        else:
            store.ui_status = 'pending'

    # 6. Dashboard Counters (Cached if possible, but distinct count is okay for now)
    # Note: These counts are for the WHOLE dataset, not just the page/filter
    # To optimize further, we could cache these or calculate only once per session
    total_stores = Store.objects.filter(active=True).count()
    verified_count = len(verified_store_ids)
    ok_count = len(ok_store_ids)
    irregular_count = len(irregular_store_ids)

    # Pendências abertas para exibição no dashboard (se o usuário for gestor/admin)
    # Adicionar paginação para pendências (10 por página)
    pending_issues = []
    pending_issues_page = None
    if request.user.role in ['gestor', 'administrador']:
        pending_issues_queryset = StoreAuditIssue.objects.filter(status='aberta').select_related('store').prefetch_related('items__audit__analyst')
        
        # Paginação para pendências (10 por página)
        from django.core.paginator import Paginator
        pending_paginator = Paginator(pending_issues_queryset, 25)
        pending_page_number = request.GET.get('pending_page', 1)
        pending_issues_page = pending_paginator.get_page(pending_page_number)
        pending_page_number = request.GET.get('pending_page', 1)
        pending_issues_page = pending_paginator.get_page(pending_page_number)
        pending_issues = pending_issues_page.object_list
        pending_issues_count = pending_paginator.count
    else:
        pending_issues_count = 0

    context = {
        'title': 'Verificação de Lojas',
        'stores': page_obj, # Now passing the Page object
        'total_stores': total_stores,
        'scope': scope, # Passed to template for filter state
        'tab': tab, # Current tab
        'filter_type': filter_type, # Ensure this is passed too
        'search_query': search_query, # Ensure this is passed too
        'verified_count': verified_count,
        'ok_count': ok_count,
        'irregular_count': irregular_count,
        'suspended_count': suspended_count,
        'filter_type': filter_type,
        'pending_issues': pending_issues,
        'pending_issues_page': pending_issues_page,  # Objeto de paginação
        'irregular_store_ids': irregular_store_ids,
        'play_sound': request.session.pop('play_irregularity_sound', False),
        'search_query': search_query,
        'pending_issues_count': pending_issues_count,
    }
    return render(request, 'core/verificacao_lojas.html', context)


@login_required
def api_store_detail(request, store_id):
    """Retorna detalhes de uma loja específica para o modal"""
    from django.http import JsonResponse
    
    try:
        store = Store.objects.get(id=store_id)
        
        # Buscar última auditoria
        last_audit = StoreAudit.objects.filter(store=store).select_related('analyst').order_by('-created_at').first()
        
        data = {
            'success': True,
            'store': {
                'id': store.id,
                'code': store.code,
                'city': store.city or 'Não informada',
                'active': store.active,
                'needs_reverification': store.needs_reverification,
                'last_audit_date': timezone.localtime(last_audit.created_at).strftime('%d/%m/%Y %H:%M') if last_audit else None,
                'last_audit_result': store.last_audit_result,
                'analyst_name': last_audit.analyst.get_full_name() or last_audit.analyst.username if last_audit else None,
            }
        }
        return JsonResponse(data)
    except Store.DoesNotExist:
        return JsonResponse({'success': False, 'error': 'Loja não encontrada'}, status=404)


@login_required
def store_audit_create(request, store_id):
    """Cria uma nova auditoria para a loja"""
    store = get_object_or_404(Store, id=store_id)
    
    if request.method == 'POST':
        # NOVO: Verificar quota diária para analistas
        if request.user.role == 'analista':
            from .models import DailyAuditQuota
            daily_quota = DailyAuditQuota.get_or_create_today(request.user)
            
            if daily_quota.is_quota_reached:
                messages.error(
                    request,
                    f'⚠️ Limite diário atingido! Você já auditou {daily_quota.audits_completed}/{daily_quota.daily_quota} lojas hoje. '
                    f'Aguarde até amanhã às 00:00 para continuar auditando.'
                )
                return redirect('verificacao_lojas')
        
        audit = StoreAudit.objects.create(analyst=request.user, store=store)
        
        items_slugs = ['cameras', 'estofados', 'cestos_medidas', 'layout', 'tv', 'totem', 'limpeza', 'marketing']
        has_irregularity = False
        
        for slug in items_slugs:
            status = request.POST.get(f'status_{slug}')
            is_compliant = (status == 'conformidade')
            photo = request.FILES.get(f'photo_{slug}')
            description = request.POST.get(f'desc_{slug}', '')
            
            # Campos específicos de Câmeras
            cameras_recording = None
            cameras_recording_mode = None
            
            if slug == 'cameras' and not is_compliant:
                # Capturar dados apenas se for câmera e estiver irregular
                rec_val = request.POST.get('cameras_recording')
                if rec_val == 'yes':
                    cameras_recording = True
                    cameras_recording_mode = request.POST.get('cameras_mode')
                elif rec_val == 'no':
                    cameras_recording = False
            
            audit_item = StoreAuditItem.objects.create(
                audit=audit,
                item_name=slug,
                is_compliant=is_compliant,
                photo=photo,
                description=description,
                cameras_recording=cameras_recording,
                cameras_recording_mode=cameras_recording_mode
            )
            
            if not is_compliant:
                has_irregularity = True
                # Busca se já existe uma pendência aberta para esta loja
                issue = StoreAuditIssue.objects.filter(store=store, status='aberta').first()
                
                if not issue:
                    # Se não existe, cria uma nova
                    issue = StoreAuditIssue.objects.create(store=store)
                
                # Vincula o item irregular à pendência (existente ou nova)
                audit_item.issue = issue
                audit_item.save()
        
        # Atualizar campos de reverificação da loja
        from django.utils import timezone
        store.last_audit_date = timezone.now()
        store.last_audit_result = 'irregular' if has_irregularity else 'conforme'
        store.needs_reverification = False  # Resetar flag de reverificação
        store.save()
        
        # NOVO: Incrementar contador de auditorias diárias
        if request.user.role == 'analista':
            from .models import DailyAuditQuota
            daily_quota = DailyAuditQuota.get_or_create_today(request.user)
            can_continue = daily_quota.increment_audits()
            
            if not can_continue:
                # Quota atingida exatamente agora
                messages.warning(
                    request,
                    f'✅ Meta diária concluída! Você auditou {daily_quota.audits_completed}/{daily_quota.daily_quota} lojas hoje. '
                    f'Parabéns! Aguarde até amanhã para continuar.'
                )
        
        if has_irregularity:
            messages.warning(request, f"Auditoria da loja {store.code} finalizada com irregularidades detectadas. Gestor notificado.")
            request.session['play_irregularity_sound'] = True # Sinal para o front tocar som
        else:
            messages.success(request, f"Auditoria da loja {store.code} finalizada com sucesso (Tudo conforme).")
            
        return redirect('verificacao_lojas')

    # Checklist items
    items_choices = [
        ('cameras', 'Câmeras'),
        ('estofados', 'Estofados'),
        ('cestos_medidas', 'Cestos de medidas'),
        ('layout', 'Layout'),
        ('tv', 'TV'),
        ('totem', 'Totem'),
        ('limpeza', 'Limpeza da loja'),
        ('marketing', 'Marketing'),
    ]

    # Histórico de auditorias da loja
    history = StoreAudit.objects.filter(store=store).order_by('-created_at')[:10]
    
    # Dados da última auditoria para exibir no formulário
    last_audit = history.first() if history else None
    last_audit_items = []
    if last_audit:
        last_audit_items = last_audit.items.filter(is_compliant=False)  # Apenas itens irregulares
    
    context = {
        'store': store,
        'history': history,
        'items_choices': items_choices,
        'title': f'Auditoria: Loja {store.code}',
        'last_audit': last_audit,
        'last_audit_items': last_audit_items,
        'needs_reverification': store.needs_reverification,
    }
    return render(request, 'core/store_audit_form.html', context)


@login_required
def store_issue_resolve(request, issue_id):
    """Resolve uma pendência de auditoria (Apenas Gestor/Admin)"""
    if not (request.user.role in ['gestor', 'administrador']):
        messages.error(request, "Apenas gestores ou administradores podem resolver pendências.")
        return redirect('verificacao_lojas')
        
    issue = get_object_or_404(StoreAuditIssue, id=issue_id)
    
    if request.method == 'POST':
        notes = request.POST.get('gestor_notes', '')
        resolution_stage = request.POST.get('resolution_stage', '')
        notification_channel = request.POST.get('notification_channel', '')
        
        # Validar que o canal de notificação foi informado
        if not notification_channel:
            store_code = issue.store.code if issue.store else 'Loja Desconhecida'
            messages.error(request, f"Por favor, informe o canal de notificação antes de resolver a pendência da loja {store_code}.")
            return redirect('verificacao_lojas')
        
        # Adicionar entrada no histórico de resolução
        import json
        history_entry = {
            'timestamp': timezone.now().isoformat(),
            'user': request.user.get_full_name() or request.user.username,
            'stage': resolution_stage,
            'channel': notification_channel,
            'notes': notes
        }
        
        # Atualizar histórico (garantir que é uma lista)
        if not isinstance(issue.resolution_history, list):
            issue.resolution_history = []
        issue.resolution_history.append(history_entry)
        
        # Atualizar campos
        issue.gestor_notes = notes
        issue.resolution_stage = 'resolvida' if resolution_stage == 'resolvida' else resolution_stage
        issue.notification_channel = notification_channel
        issue.status = 'resolvida'
        issue.resolved_at = timezone.now()
        issue.resolved_by = request.user
        issue.save()
        
        store_code = issue.store.code if issue.store else 'Loja Desconhecida'
        messages.success(request, f"Pendência em {store_code} marcada como resolvida.")
        
    return redirect('verificacao_lojas')


@login_required
def store_create(request):
    """Cria uma nova loja (Apenas Gestor/Admin)"""
    if request.user.role not in ['gestor', 'administrador']:
        messages.error(request, "Você não tem permissão para criar lojas.")
        return redirect('verificacao_lojas')

    if request.method == 'POST':
        form = StoreForm(request.POST)
        if form.is_valid():
            form.save()
            messages.success(request, "Loja criada com sucesso!")
            return redirect('verificacao_lojas')
    else:
        form = StoreForm()

    return render(request, 'core/store_form.html', {
        'form': form,
        'title': 'Cadastrar Nova Loja'
    })


@login_required
def store_edit(request, store_id):
    """Edita uma loja existente (Apenas Gestor/Admin)"""
    if request.user.role not in ['gestor', 'administrador']:
        messages.error(request, "Apenas gestores ou administradores podem editar lojas.")
        return redirect('verificacao_lojas')

    store = get_object_or_404(Store, id=store_id)

    if request.method == 'POST':
        form = StoreForm(request.POST, instance=store)
        if form.is_valid():
            form.save()
            messages.success(request, f"Loja {store.code} atualizada com sucesso!")
            return redirect('verificacao_lojas')
    else:
        form = StoreForm(instance=store)

    return render(request, 'core/store_form.html', {
        'form': form,
        'title': f'Editar Loja: {store.code}'
    })


@login_required
def store_delete(request, store_id):
    """Exclui uma loja (Apenas Gestor/Admin)"""
    if request.user.role not in ['gestor', 'administrador']:
        messages.error(request, "Apenas gestores ou administradores podem excluir lojas.")
        return redirect('verificacao_lojas')

    store = get_object_or_404(Store, id=store_id)

    if request.method == 'POST':
        code = store.code
        store.delete()
        messages.success(request, f"Loja {code} excluída com sucesso.")
        return redirect('verificacao_lojas')

    return render(request, 'core/confirm_delete.html', {
        'title': f'Excluir Loja {store.code}',
        'message': f'Tem certeza que deseja excluir a loja {store.code}? Todas as auditorias e pendências associadas também serão removidas.',
        'back_url': 'verificacao_lojas'
    })


@login_required
def store_bulk_delete(request):
    """Exclui TODAS as lojas (Apenas Administrador)"""
    if not request.user.is_administrador():
        messages.error(request, "Acesso negado. Apenas administradores podem realizar esta ação.")
        return redirect('verificacao_lojas')

    if request.method == 'POST':
        count = Store.objects.all().count()
        # Delete all stores (cascades to audits/issues)
        Store.objects.all().delete()
        messages.success(request, f"Sucesso! {count} lojas foram excluídas permanentemente.")
        return redirect('verificacao_lojas')

    return render(request, 'core/confirm_delete.html', {
        'title': 'EXCLUIR TODAS AS LOJAS',
        'message': 'ATENÇÃO: Você está prestes a excluir TODAS as lojas cadastradas, incluindo todas as auditorias e históricos. Esta ação é irreversível. Tem certeza absoluta?',
        'back_url': 'verificacao_lojas',
        'confirm_btn_class': 'btn-danger' # Optional if supported by template
    })


@login_required
def store_issue_edit(request, issue_id):
    """Edita uma pendência de auditoria (Apenas Gestor/Admin)"""
    if request.user.role not in ['gestor', 'administrador']:
        messages.error(request, "Permissão negada.")
        return redirect('verificacao_lojas')

    issue = get_object_or_404(StoreAuditIssue, id=issue_id)
    
    if request.method == 'POST':
        new_status = request.POST.get('status')
        new_notes = request.POST.get('gestor_notes')
        
        if new_status:
            issue.status = new_status
        if new_notes is not None:
             issue.gestor_notes = new_notes
             
        issue.save()
        messages.success(request, "Pendência atualizada com sucesso.")
        return redirect('verificacao_lojas')

    # Como é uma edição simples, reutilizamos o template de confirmação/edição ou criamos um simples
    # Vou usar um template simples dedicado ou reusar o form genérico
    store_code = issue.store.code if issue.store else 'L000 (Desconhecida)'

    return render(request, 'core/issue_edit_form.html', {
        'issue': issue,
        'title': f'Editar Pendência: {store_code}'
    })


@login_required
def store_issue_delete(request, issue_id):
    """Exclui uma pendência de auditoria (Apenas Gestor/Admin)"""
    if request.user.role not in ['gestor', 'administrador']:
        messages.error(request, "Permissão negada.")
        return redirect('verificacao_lojas')

    issue = get_object_or_404(StoreAuditIssue, id=issue_id)

    if request.method == 'POST':
        issue.delete()
        messages.success(request, "Pendência excluída com sucesso.")
        return redirect('verificacao_lojas')

    store_code = issue.store.code if issue.store else 'L000 (Desconhecida)'

    return render(request, 'core/confirm_delete.html', {
        'title': 'Excluir Pendência',
        'message': f'Tem certeza que deseja excluir a pendência da loja {store_code}? Isso também excluirá os itens irregulares associados.',
        'back_url': 'verificacao_lojas'
    })


@login_required
def import_stores_xlsx(request):
    """Importa base de lojas via XLSX (Apenas Gestor/Admin)"""
    if request.user.role not in ['gestor', 'administrador']:
        messages.error(request, "Você não tem permissão para importar lojas.")
        return redirect('verificacao_lojas')

    if request.method == 'POST':
        if 'xlsx_file' not in request.FILES:
            messages.error(request, "Selecione um arquivo XLSX.")
            return redirect('import_stores_xlsx')

        file = request.FILES['xlsx_file']
        try:
            from openpyxl import load_workbook
            wb = load_workbook(file, data_only=True)
            ws = wb.active
            
            created = 0
            updated = 0
            
            for row in ws.iter_rows(min_row=2, values_only=True):
                code = str(row[0]).strip().upper() if row and row[0] else None
                
                if code:
                    store, is_created = Store.objects.get_or_create(
                        code=code,
                        defaults={'active': True}
                    )
                    if is_created: created += 1
                    else: updated += 1
            
            messages.success(request, f"Importação concluída: {created} criadas, {updated} atualizadas.")
            return redirect('verificacao_lojas')
        except Exception as e:
            messages.error(request, f"Erro ao processar arquivo: {str(e)}")
            return redirect('import_stores_xlsx')

    return render(request, 'core/import_stores.html', {'title': 'Importar Lojas (XLSX)'})


# ========================================
# CHAT
# ========================================




# ========================================
# AUDITORIA DE ATENDIMENTOS
# ========================================

@login_required
def auditoria_atendimentos_view(request):
    """Página de Auditoria de Atendimentos - Gestores, Admins e Analistas"""
    if not (request.user.is_gestor() or request.user.is_administrador() or request.user.is_analista()):
        messages.error(request, "Acesso negado. Você não tem permissão para acessar esta página.")
        return redirect('dashboard')
    
    # Obter departamento
    department = request.session.get('current_department_obj') or request.user.department
    
    context = {
        'title': 'Auditoria de Atendimentos',
        'department': department,
        'is_admin': request.user.is_administrador(),
        'is_gestor': request.user.is_gestor(),
        'is_analista': request.user.is_analista(),
    }
    return render(request, 'core/auditoria_atendimentos.html', context)

@login_required
def api_get_system_notifications(request):
    """
    API para retornar as últimas notificações do sistema.
    Retorna JSON com lista de notificações ativas.
    """
    notifications = SystemNotification.objects.filter(is_active=True).order_by('-created_at')[:5]
    
    data = []
    for notification in notifications:
        data.append({
            'id': notification.id,
            'title': notification.title,
            'message': notification.message,
            'details': notification.details,
            'category': notification.get_category_display(),
            'category_code': notification.category,
            'created_at': notification.created_at.strftime('%d/%m/%Y %H:%M'),
            'timestamp': notification.created_at.timestamp()
        })
        
    return JsonResponse({'notifications': data})
